"""
core/hazmat_exemption.py
========================
Prepare Amazon's FBA dangerous-goods EXEMPTION SHEETS from an ASIN's SP-API data.

Picks the right template (battery vs no-harmful-chemicals), fills the safe
non-dropdown fields (ASIN, title, what's-in-the-box), and surfaces detected battery
details as HINTS. The cascading dropdowns (Batteries sold → Composition → Cells →
Watt-hours) are left for the seller to complete in Excel — they're linked named
ranges, so filling them blind risks rejection. Matches the 'review then upload' flow.
"""
from __future__ import annotations
import io
import os
import re

_HERE = os.path.dirname(os.path.abspath(__file__))
_TPL = os.path.join(_HERE, "..", "data", "hazmat_templates")
BATTERY_TPL = os.path.join(_TPL, "battery.xlsx")
CHEMICAL_TPL = os.path.join(_TPL, "chemical.xlsx")

# Seller signature for the exemption sheet (cells D25 / G25).
SIGNER_FIRST = "Nima"
SIGNER_LAST = "Zare"

_BATTERY_HINTS = ("batter", "lithium", "cell", "watt", "voltage", "dg_hz")
_BATTERY_WORDS = ("battery", "power bank", "powerbank", "charger", "rechargeable",
                  "lithium", "fan", "speaker", "headphone", "earbud", "jump starter",
                  "power station", "massager", "trimmer", "shaver", "light", "torch")


def _attr_val(attrs: dict, key: str):
    v = attrs.get(key)
    if isinstance(v, list) and v:
        e = v[0]
        return e.get("value", e.get("name", "")) if isinstance(e, dict) else e
    return v


def classify(item: dict) -> str:
    """'battery' if the item has battery attributes or is a battery-powered type;
    otherwise 'chemical' (no-harmful-chemicals)."""
    attrs = item.get("attributes") or {}
    if any(any(h in k.lower() for h in _BATTERY_HINTS) for k in attrs):
        return "battery"
    blob = f"{item.get('title','')} {item.get('product_type','')}".lower()
    return "battery" if any(w in blob for w in _BATTERY_WORDS) else "chemical"


def _map_composition(detected: str, title: str) -> str:
    d, t = (detected or "").lower(), (title or "").lower()
    if "polymer" in d or "li-po" in t or "lipo" in t:
        return "Lithium_Polymer"
    if "metal" in d:
        return "Lithium_Metal"
    if "iron" in d or "lifepo" in d or "phosphate" in d:
        return "Lithium_iron_phosphate"
    if "nmc" in d or "manganese" in d:
        return "Lithium_nickel_manganese_cobalt_oxide"
    if "cobalt" in d:
        return "Lithium_cobalt_oxide"
    if "titanate" in d:
        return "Lithium_titanate"
    if "alkaline" in d:
        return "Alkaline"
    if "lead" in d:
        return "Lead_Acid"
    return "Lithium_Ion"            # default for lithium-powered electronics


def _map_packaging(detected: str) -> str:
    d = (detected or "").lower()
    if "packed" in d:
        return "Packed with"
    if "only" in d or "stand" in d:
        return "Stand alone"
    return "In Equipment"            # contained in equipment (matches sample)


def _voltage(item: dict) -> float:
    """Nominal battery voltage. Use the battery-voltage attribute if present; else the
    lithium nominal 3.7 V. (We do NOT read 'XV' from the title — that's usually the
    OUTPUT voltage, not the cell's nominal voltage, and would inflate Wh.)"""
    attrs = item.get("attributes") or {}
    v = _attr_val(attrs, "battery_voltage") or _attr_val(attrs, "nominal_voltage")
    try:
        v = float(v)
        if 0 < v < 60:
            return v
    except (TypeError, ValueError):
        pass
    return 3.7


def _watt_hours(item: dict):
    """Battery energy in Wh, or None. Wh = capacity(Ah) × voltage(V). Capacity comes
    from the lithium_battery attribute, else from 'NNNNmAh' in the title."""
    attrs = item.get("attributes") or {}
    lb = attrs.get("lithium_battery")
    if isinstance(lb, list) and lb and isinstance(lb[0], dict):
        ec = lb[0].get("energy_content")
        if isinstance(ec, list) and ec and isinstance(ec[0], dict):
            try:
                val = float(ec[0].get("value"))
                unit = str(ec[0].get("unit", "")).lower()
                if "watt" in unit:
                    return val
                if "ampere" in unit or "mah" in unit:
                    return val / 1000.0 * _voltage(item)
            except (TypeError, ValueError):
                pass
    title = item.get("title", "") or ""
    m = re.search(r"(\d[\d,]*)\s*mah", title, re.I)            # e.g. "10000mAh"
    if m:
        try:
            return float(m.group(1).replace(",", "")) / 1000.0 * _voltage(item)
        except ValueError:
            pass
    m = re.search(r"(\d+(?:\.\d+)?)\s*wh\b", title, re.I)       # e.g. "37Wh"
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            pass
    return None


def _wh_bucket(wh, single_cell: bool) -> str:
    """Bucket a Wh value into the sheet's exact dropdown ranges."""
    if wh is None:
        return ""
    if single_cell:
        return "WH <= 20" if wh <= 20 else "WH > 20"
    if wh <= 100:
        return "WH <= 100"
    return "101 - 300 WH" if wh <= 300 else "WH > 300"


def _whats_in_box(item: dict) -> str:
    pt = item.get("product_type", "")
    if pt:
        return pt.replace("_", " ").title().replace("Usb", "USB")
    return (item.get("title", "") or "Item")[:60]


def battery_fields(item: dict) -> dict:
    """Derive the battery declaration values (exact dropdown strings) from SP-API data."""
    attrs = item.get("attributes") or {}
    cells_raw = (_attr_val(attrs, "number_of_lithium_ion_cells")
                 or _attr_val(attrs, "number_of_cells")
                 or _attr_val(attrs, "number_of_lithium_metal_cells") or 1)
    try:
        single = int(float(cells_raw)) <= 1
    except (TypeError, ValueError):
        single = True
    comp = _map_composition(_attr_val(attrs, "battery_cell_composition")
                            or _attr_val(attrs, "compliance_battery_chemical_construction"),
                            item.get("title"))
    wh = _watt_hours(item)
    return {
        "whats_in_box": _whats_in_box(item),
        "batteries_sold": "Yes",
        "composition": comp,
        "packaging": _map_packaging(_attr_val(attrs, "lithium_battery_packaging")),
        "cells": "Single_cell" if single else "Multiple_cells",
        "watt_hours": _wh_bucket(wh, single),
        "wh_value": round(wh, 1) if wh is not None else None,
    }


def _set(ws, coord, value):
    """Write a value, tolerating merged cells (write to the merge's top-left)."""
    try:
        ws[coord] = value
    except Exception:
        for rng in ws.merged_cells.ranges:
            if coord in rng:
                ws.cell(rng.min_row, rng.min_col).value = value
                return


def fill(item: dict, kind: str) -> bytes:
    """Return the chosen template (bytes) with the safe fields pre-filled."""
    from openpyxl import load_workbook
    if kind == "battery":
        wb = load_workbook(BATTERY_TPL)
        ws = wb["Battery exemption sheet"]
        f = battery_fields(item)
        _set(ws, "I13", item.get("asin", ""))                       # ASIN
        _set(ws, "J13", (item.get("title", "") or "")[:250])        # Product title
        _set(ws, "K13", f["whats_in_box"])                          # What's in the box
        _set(ws, "L13", f["batteries_sold"])                        # Batteries sold / is a battery
        _set(ws, "M13", f["composition"])                           # Chemical composition / cell type
        _set(ws, "N13", f["packaging"])                             # Battery packaging
        _set(ws, "O13", f["cells"])                                 # No. of cells
        if f["watt_hours"]:
            _set(ws, "P13", f["watt_hours"])                        # Watt-hours range
        _set(ws, "D25", SIGNER_FIRST)                               # signature: first name
        _set(ws, "G25", SIGNER_LAST)                                # signature: last name
    else:
        wb = load_workbook(CHEMICAL_TPL)
        ws = wb["No harmful chemicals"]
        _set(ws, "N17", item.get("asin", ""))                       # ASIN
        _set(ws, "O17", (item.get("title", "") or "")[:250])        # Product name
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
