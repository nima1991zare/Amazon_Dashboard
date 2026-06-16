"""
modules/hazmat_inactive.py
==========================
☣️ Hazmat & Inactive (new module). Two parts:

PART 1 — HAZMAT (FBA Dangerous Goods compliance)
  Monitors the FBA Compliance Dashboard (via api_client; mock now, real API later
  — sellercentral.amazon.ae/fba/compliance-dashboard). Maps each item's status:
    * "Unable to classify"            -> generate a Battery Exemption sheet and
                                         upload it (button generates an upload-ready
                                         .xlsx using Amazon's exact template).
    * "Dangerous Good FBA Fulfillable"-> APPROVED -> set fulfilment channel to FBA.
    * "Dangerous Good Unfulfillable"  -> NOT approved -> set fulfilment channel FBM.
  Actions call api_client stubs and log tasks.

PART 2 — INACTIVE
  Lists inactive listings with the reason, Export, and one-click tasks to fix.

You can also upload the dashboard's own export (CSV/Excel) instead of mock data.
"""

from __future__ import annotations
import io
import re
import pandas as pd
import streamlit as st

from core import db, mock_data, compliance_scraper, sp_api, hazmat_exemption
from core.api_client import client
from core.util import get_field
from core.components import styled_table, export_buttons, kpi_row, page_header
from core.styles import section_label, badge, alert


def _exemption_prep_ui() -> None:
    """Two ASIN inputs → fetch item → pick battery/chemical template → fill safe
    fields → review + download (then upload in Seller Central)."""
    st.markdown(section_label("Prepare exemption sheet — enter an ASIN"),
                unsafe_allow_html=True)
    st.caption("Enter an ASIN → it fetches the item from Amazon, picks the right template "
               "(battery vs no-harmful-chemicals), fills ASIN / title / what's-in-the-box, and "
               "shows detected battery details. Review, complete the guided dropdowns in Excel, "
               "then upload it in Seller Central.")
    cols = st.columns(2)
    for i, col in enumerate(cols):
        with col:
            asin = st.text_input(f"ASIN {i + 1}", key=f"hz_ex_asin_{i}", placeholder="B0…")
            if st.button(f"Prepare sheet {i + 1}", key=f"hz_ex_btn_{i}",
                         use_container_width=True) and asin.strip():
                with st.spinner("Fetching item from Amazon…"):
                    item = sp_api.get_item_by_asin(asin.strip())
                if not item.get("ok"):
                    st.session_state.pop(f"hz_ex_{i}", None)
                    st.error(f"Lookup failed: {item.get('reason')}")
                else:
                    kind = hazmat_exemption.classify(item)
                    st.session_state[f"hz_ex_{i}"] = {
                        "item": item, "kind": kind,
                        "bytes": hazmat_exemption.fill(item, kind),
                        "fields": hazmat_exemption.battery_fields(item) if kind == "battery" else {}}
            data = st.session_state.get(f"hz_ex_{i}")
            if data:
                it, kind = data["item"], data["kind"]
                st.markdown(badge("🔋 Battery exemption" if kind == "battery"
                                  else "🧪 No-harmful-chemicals exemption",
                                  "blue" if kind == "battery" else "violet"),
                            unsafe_allow_html=True)
                st.caption(f"**{(it.get('title', '') or '')[:90]}**  \n"
                           f"Brand: {it.get('brand', '') or '—'} · Type: {it.get('product_type', '') or '—'}")
                if kind == "battery" and data["fields"]:
                    f = data["fields"]
                    wh = f" (~{f['wh_value']} Wh)" if f.get("wh_value") else ""
                    st.caption(f"Auto-filled → batteries sold: **{f['batteries_sold']}** · "
                               f"composition: **{f['composition']}** · packaging: **{f['packaging']}** · "
                               f"cells: **{f['cells']}** · watt-hours: "
                               f"**{f['watt_hours'] or '— set in sheet'}**{wh}. "
                               f"Signed **{hazmat_exemption.SIGNER_FIRST} "
                               f"{hazmat_exemption.SIGNER_LAST}**. Review & upload.")
                st.download_button(
                    "⬇ Download filled sheet (review)", data["bytes"],
                    file_name=f"exemption_{kind}_{it.get('asin', '')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"hz_ex_dl_{i}", use_container_width=True)
                if st.button("⬆ Upload to Amazon", key=f"hz_ex_up_{i}",
                             use_container_width=True, type="primary"):
                    import os
                    asin = it.get("asin", "item")
                    tmp = os.path.abspath(os.path.join("data", f".upload_{asin}.xlsx"))
                    with open(tmp, "wb") as fh:
                        fh.write(data["bytes"])
                    ok, msg = compliance_scraper.upload(asin, tmp)
                    (st.success if ok else st.error)(msg)
                st.caption("Upload selects **English** and submits automatically. When it's "
                           "done, use **🔄 Pull compliance data** above to refresh the table.")


def _fmt_when(iso: str) -> str:
    from datetime import datetime
    try:
        return datetime.fromisoformat(str(iso)).astimezone().strftime("%Y-%m-%d %H:%M")
    except Exception:
        return str(iso or "")


def _hazmat_changes_section() -> None:
    """Show which items changed classification since prior pulls (persisted in the DB)."""
    st.markdown(section_label("🔔 Status changes"), unsafe_allow_html=True)
    last = st.session_state.get("hz_changes")
    if last is not None:
        st.caption(f"Last pull detected **{len(last)}** status change(s).")
    changes = db.get_hazmat_changes(200)
    if not changes:
        st.caption("No status changes recorded yet — pull again later and any item whose "
                   "classification changed will appear here.")
        return
    dfc = pd.DataFrame([{
        "Changed": _fmt_when(c["changed_at"]), "ASIN": c["asin"],
        "Title": (c["title"] or "")[:48],
        "Was": c["old_status"] or "—", "Now": c["new_status"] or "—"} for c in changes])
    styled_table(dfc, highlight={"row-warn": lambda r: True})
    export_buttons(dfc, "hazmat_status_changes")


def _norm_rows(rows: list) -> list:
    """Normalize scraped/loaded dict rows (varying keys) → the columns the tab uses."""
    def pick(d, *names, default=""):
        nm = {re.sub(r"[^a-z0-9]", "", str(k).lower()): k for k in d.keys()}
        for name in names:
            n = re.sub(r"[^a-z0-9]", "", name.lower())
            if n in nm:
                return str(d[nm[n]])
            for cn, orig in nm.items():
                if n and n in cn:
                    return str(d[orig])
        return default
    out = []
    for d in rows or []:
        out.append({
            "asin": pick(d, "asin"),
            "sku": pick(d, "msku", "sellersku", "sku"),
            "title": pick(d, "itemname", "productname", "productdescription", "title", "product"),
            "hazmat_status": pick(d, "fulfillmentclassification", "dangerousgoodsclassification",
                                  "classificationstatus", "classification", "disposition", "status"),
            "fulfilment_channel": pick(d, "fulfillmentchannel", "channel", default="FBM"),
        })
    return out


# ---------------------------------------------------------------------------
# Status classification
# ---------------------------------------------------------------------------
def _hazmat_bucket(status: str) -> str:
    s = str(status).strip().lower()
    if not s or s == "nan":
        return "other"
    if any(w in s for w in ("unable", "pending", "unclassified", "not classified",
                            "in progress", "in review", "requires")):
        return "unable"
    # not-fulfillable / dangerous (check before the generic 'fulfil' match!)
    if any(w in s for w in ("not fulfil", "unfulfil", "prohibited", "restricted",
                            "not eligible", "blocked", "removed")):
        return "unfulfillable"
    if any(w in s for w in ("not a dangerous", "not regulated", "not hazmat", "exempt",
                            "approved", "fulfil", "eligible", "compliant", "classified")):
        return "fulfillable"   # 'classified' = Amazon has a valid classification on file
    return "other"


def _col(raw: pd.DataFrame, *names, default: str = ""):
    """Return the first column matching any of `names` (by normalized header,
    exact then substring) as a string Series; else a column of `default`."""
    def n(c):
        return re.sub(r"[^a-z0-9]", "", str(c).lower())
    norm = {n(c): c for c in raw.columns}
    for name in names:
        nm = re.sub(r"[^a-z0-9]", "", name.lower())
        if nm in norm:
            return raw[norm[nm]].fillna("").astype(str)
        for cn, orig in norm.items():
            if nm and nm in cn:
                return raw[orig].fillna("").astype(str)
    return pd.Series([default] * len(raw))


_ACTION = {
    "unable":        "Generate & upload hazmat exemption sheet",
    "fulfillable":   "Approved → set fulfilment channel to FBA",
    "unfulfillable": "Not approved → set fulfilment channel to FBM",
    "other":         "Review manually",
}
_BADGE = {
    "unable": ("⚠ Unable to classify", "amber"),
    "fulfillable": ("✓ FBA Fulfillable (approved)", "green"),
    "unfulfillable": ("✗ Unfulfillable", "coral"),
    "other": ("Review", "violet"),
}


# ---------------------------------------------------------------------------
# Battery exemption sheet generator (Amazon's exact template + dropdowns)
# ---------------------------------------------------------------------------
def build_battery_exemption_xlsx(items: pd.DataFrame) -> bytes:
    """Return an upload-ready .xlsx pre-filled with ASIN + title for the given
    items, with Amazon's exact headers and data-validation dropdowns on the
    battery columns. The remaining cells are left for the user/operator to fill.
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "Battery exemption sheet"
    ws.append(mock_data.HAZMAT_TEMPLATE_HEADERS)
    for _, r in items.iterrows():
        ws.append([r.get("asin", ""), r.get("title", ""), "", "", "", "", "", "", "", ""])

    # Hidden sheet holding the allowed dropdown values.
    lists = wb.create_sheet("Lists")
    a = mock_data.HAZMAT_ALLOWED
    columns = {"A": a["sold"], "B": a["chemical"], "C": a["packaging"],
               "D": a["cells"], "E": a["watt_hours"], "F": a["spillability"]}
    for col, vals in columns.items():
        for i, v in enumerate(vals, start=1):
            lists[f"{col}{i}"] = v
    lists.sheet_state = "hidden"

    last_row = ws.max_row + 100  # allow extra manual rows
    # (sheet column letter, lists column, value count)
    dv_map = [("D", "A", len(a["sold"])), ("E", "B", len(a["chemical"])),
              ("F", "C", len(a["packaging"])), ("G", "D", len(a["cells"])),
              ("H", "E", len(a["watt_hours"])), ("I", "F", len(a["spillability"]))]
    for sheet_col, list_col, n in dv_map:
        dv = DataValidation(type="list",
                            formula1=f"Lists!${list_col}$1:${list_col}${n}",
                            allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{sheet_col}2:{sheet_col}{last_row}")

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


# ---------------------------------------------------------------------------
# PART 1 — Hazmat
# ---------------------------------------------------------------------------
def _hazmat_tab() -> None:
    st.markdown(
        "<p style='color:var(--muted); font-size:.86rem'>Real data from the "
        "<b>FBA Compliance Dashboard</b>. Amazon has no API for it, so this drives a logged-in "
        "browser to read the dashboard for you (or upload its export).</p>",
        unsafe_allow_html=True)

    # --- Auto-pull from Seller Central (browser) --------------------------
    st.markdown(section_label("Auto-pull from Seller Central (browser)"), unsafe_allow_html=True)
    if not compliance_scraper.available():
        st.warning("Playwright isn't installed. In the project folder run:  "
                   "`python -m pip install playwright`  then  `python -m playwright install chromium` "
                   "(or just have Google Chrome installed).")
    else:
        cc = st.columns([1, 1, 2])
        if cc[0].button("🔓 Log in to Seller Central", use_container_width=True):
            ok, msg = compliance_scraper.login()
            (st.success if ok else st.error)(msg)
        if cc[1].button("🔄 Pull compliance data", use_container_width=True, type="primary"):
            with st.spinner("Opening Seller Central and reading the compliance dashboard…"):
                res = compliance_scraper.fetch()
            if res.get("ok"):
                rows = _norm_rows(res.get("rows", []))
                for r in rows:                              # tag bucket + clean status
                    r["bucket"] = _hazmat_bucket(r.get("hazmat_status", ""))
                    r["status"] = (r.get("hazmat_status", "") or "").split("\n")[0].strip()
                st.session_state["hz_changes"] = db.save_hazmat_statuses(rows)
                st.session_state["hz_scraped"] = rows
                ch = len(st.session_state["hz_changes"])
                st.success(f"Pulled {len(rows)} item(s) — {ch} status change(s) since last pull.")
                st.rerun()
            else:
                st.error(f"Couldn't pull: {res.get('reason')}")
        cc[2].caption("First time: click **Log in**, sign in (incl. OTP), close that window — "
                      "then click **Pull**. Pull reads the **last 10 pages** of the dashboard "
                      "(~1–2 min); the session is remembered for next time.")

    _exemption_prep_ui()

    up = st.file_uploader("…or upload the dashboard export (CSV/Excel)",
                          type=["csv", "xlsx"], key="hz_up")

    # --- Resolve data source: scraped (preferred) → uploaded → nothing ----
    scraped = st.session_state.get("hz_scraped")
    if scraped:
        df = pd.DataFrame(scraped)
        src = f"Seller Central (browser) — {len(df)} items"
    elif up:
        raw = pd.read_csv(up) if up.name.endswith(".csv") else pd.read_excel(up)
        df = pd.DataFrame(_norm_rows(raw.to_dict("records")))
        src = f"{up.name} — {len(df)} items"
    else:
        st.info("⬆ Use **Auto-pull** above, or upload your dashboard export, to load real data.")
        return
    st.markdown(badge(f"Source: {src} (real)", "green"), unsafe_allow_html=True)

    df = df.copy()
    df["bucket"] = df["hazmat_status"].map(_hazmat_bucket)
    overrides = db.get_channel_overrides()
    df["fulfilment_channel"] = df.apply(
        lambda r: overrides.get(r["sku"], r["fulfilment_channel"]), axis=1)

    n_unable = int((df["bucket"] == "unable").sum())
    n_fba = int((df["bucket"] == "fulfillable").sum())
    n_fbm = int((df["bucket"] == "unfulfillable").sum())

    # --- Card box ---------------------------------------------------------
    kpi_row([
        {"label": "Unable to classify", "value": str(n_unable), "accent": "amber",
         "sub": "Need exemption sheet"},
        {"label": "FBA Fulfillable", "value": str(n_fba), "accent": "emerald",
         "sub": "Approved → FBA"},
        {"label": "Unfulfillable", "value": str(n_fbm), "accent": "coral",
         "sub": "Dangerous → FBM"},
    ])
    st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

    # --- Status changes (remembered across pulls) -------------------------
    _hazmat_changes_section()

    # --- Chart (green = fulfillable, yellow = unable, red = dangerous) -----
    import altair as alt
    chart_df = pd.DataFrame({
        "Status": ["FBA Fulfillable", "Unable to classify", "Dangerous (Unfulfillable)"],
        "Count": [n_fba, n_unable, n_fbm]})
    scale = alt.Scale(
        domain=["FBA Fulfillable", "Unable to classify", "Dangerous (Unfulfillable)"],
        range=["#2e9e5b", "#f6c343", "#e7503c"])   # green / yellow / red
    chart = (alt.Chart(chart_df).mark_bar(cornerRadiusTopLeft=4, cornerRadiusTopRight=4)
             .encode(
                 x=alt.X("Status:N", sort=None, axis=alt.Axis(labelAngle=0, title=None)),
                 y=alt.Y("Count:Q", title="Items"),
                 color=alt.Color("Status:N", scale=scale, legend=None),
                 tooltip=["Status", "Count"])
             .properties(height=280))
    st.altair_chart(chart, use_container_width=True)
    st.markdown("<div style='height:6px'></div>", unsafe_allow_html=True)

    # --- Table: ASIN / Title / Fulfillment classification status ----------
    # green = FBA Fulfillable (approved), yellow = Unable to classify,
    # red = Dangerous Good (Unfulfillable).
    st.markdown(section_label("Hazmat classification"), unsafe_allow_html=True)
    show = pd.DataFrame({
        "ASIN": df["asin"],
        "Title": df["title"],
        "Fulfillment classification": df["bucket"].map(lambda b: _BADGE[b][0]),
    })
    styled_table(
        show,
        highlight={"row-danger": lambda r: "Unfulfillable" in r["Fulfillment classification"],
                   "row-warn": lambda r: "Unable" in r["Fulfillment classification"],
                   "row-good": lambda r: "Fulfillable" in r["Fulfillment classification"]},
        badge_cols={"Fulfillment classification": {v[0]: v for v in _BADGE.values()}})

    # --- Export -----------------------------------------------------------
    export_buttons(df.drop(columns=["bucket"]), "hazmat_compliance")


# ---------------------------------------------------------------------------
# PART 2 — Inactive
# ---------------------------------------------------------------------------
def _inactive_tab() -> None:
    up = st.file_uploader("Upload inactive listings export (CSV/Excel) — optional",
                          type=["csv", "xlsx"], key="inact_up")
    if up:
        raw = pd.read_csv(up) if up.name.endswith(".csv") else pd.read_excel(up)
        df = pd.DataFrame({
            "asin": get_field(raw, "asin", ""), "sku": get_field(raw, "sku", ""),
            "title": get_field(raw, "title", ""), "status": "Inactive",
            "reason": get_field(raw, "category", "Unknown"),
            "last_active": get_field(raw, "category", ""),
        })
        src = f"uploaded ({up.name})"
    else:
        df = client().get_inactive_listings()
        src = "sample data (upload to override)"
    st.markdown(badge(f"Source: {src}", "blue"), unsafe_allow_html=True)

    kpi_row([
        {"label": "Inactive listings", "value": str(len(df)), "accent": "coral"},
        {"label": "Out of stock",
         "value": str(int(df["reason"].str.contains("stock", case=False, na=False).sum())),
         "accent": "amber"},
        {"label": "Suppressed",
         "value": str(int(df["reason"].str.contains("suppress", case=False, na=False).sum())),
         "accent": "violet"},
    ])
    st.markdown("<div style='height:14px'></div>", unsafe_allow_html=True)

    st.markdown(section_label("Inactive Listings"), unsafe_allow_html=True)
    styled_table(df, highlight={"row-danger": lambda r: True})
    export_buttons(df, "inactive_listings")

    if st.button("➕ Add reactivation tasks"):
        for _, r in df.iterrows():
            db.add_task(f"Reactivate: {r['title']}", f"Inactive — {r['reason']}.",
                        module="Hazmat & Inactive", priority="medium", related_id=r["sku"])
        st.success("Reactivation tasks added.")


def render(nav=None) -> None:
    page_header("Hazmat & Inactive",
                "Dangerous-goods compliance and inactive-listing recovery", icon="☣️")
    t1, t2 = st.tabs(["☣️ Hazmat (Dangerous Goods)", "💤 Inactive"])
    with t1:
        _hazmat_tab()
    with t2:
        _inactive_tab()
