"""
modules/intake.py
=================
📥 Inventory & Listing Intake + Auto Item Creation (Module 2).

Sub-tabs:
  1. New Arrivals    — warehouse items not in your Amazon catalog (Mark as listed)
  2. Restock Pending — catalogued items needing replenishment
  3. Auto Item Creation:
        MODE A (price list): upload CSV/Excel w/ media-link column → fetch images,
                              take price from file.
        MODE B (URL): paste item URLs → generic scrape (title/price/specs/images).
        Both → build draft → optimize copy → REVIEW TABLE (editable) → Approve →
        "Ready to List" queue (persisted in DB), with Export.

Fetching is routed through oskar_source (USE_MOCK now). Listing copy via the
optimization module. Flagged rows (no price / no image) get an amber badge.
"""

from __future__ import annotations
import re
import io
import json
import time
import pandas as pd
import streamlit as st

from core import db, inventory_source, oskar_source
from core.api_client import client
from core.util import get_field, find_column
from core.components import styled_table, export_buttons, page_header
from core.styles import section_label, badge, alert
from modules import optimization


def _classify(inv: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Split inventory into new arrivals vs restock by comparing to the catalog."""
    catalog = db.get_catalog_skus()
    inv = inv.copy()
    new = inv[~inv["sku"].isin(catalog)].copy()
    restock = inv[inv["sku"].isin(catalog)].copy()
    return new.reset_index(drop=True), restock.reset_index(drop=True)


def _new_arrivals_tab() -> None:
    st.markdown(section_label("Fetch from inventory website (or upload below)"),
                unsafe_allow_html=True)
    up = st.file_uploader("Manual fallback: upload product list (CSV/Excel)",
                          type=["csv", "xlsx"], key="intake_upload")

    if up:
        inv = pd.read_csv(up) if up.name.endswith(".csv") else pd.read_excel(up)
        # Normalize to expected columns via fuzzy detection.
        inv = pd.DataFrame({
            "sku": get_field(inv, "sku", ""), "title": get_field(inv, "title", ""),
            "brand": get_field(inv, "brand", ""), "category": get_field(inv, "category", ""),
            "warehouse_qty": get_field(inv, "qty", 0), "cost": get_field(inv, "price", 0),
        })
        src = f"uploaded file ({up.name})"
    else:
        inv, src = inventory_source.fetch_inventory()

    st.markdown(badge(f"Source: {src}", "blue"), unsafe_allow_html=True)
    if inv.empty:
        st.warning("No inventory returned.")
        return

    new, _ = _classify(inv)
    st.markdown(f"<p style='color:var(--muted)'>{len(new)} items not yet in your Amazon "
                f"catalog.</p>", unsafe_allow_html=True)

    # Optional brand filter.
    brands = ["All"] + sorted(new["brand"].dropna().unique().tolist())
    pick = st.selectbox("Filter by brand", brands, key="na_brand")
    view = new if pick == "All" else new[new["brand"] == pick]

    styled_table(view, highlight={"row-good": lambda r: True})
    export_buttons(view, "new_arrivals")

    st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
    cols = st.columns([3, 1])
    with cols[0]:
        to_list = st.multiselect("Select SKUs to mark as listed", view["sku"].tolist(),
                                 key="na_marklist")
    with cols[1]:
        st.markdown("<div style='height:28px'></div>", unsafe_allow_html=True)
        if st.button("✅ Mark as Listed", use_container_width=True) and to_list:
            for sku in to_list:
                row = view[view["sku"] == sku].iloc[0]
                db.upsert_catalog_item(sku=sku, title=row["title"], brand=row["brand"],
                                       category=row["category"], status="listed")
                db.add_task(f"Verify new listing live: {row['title']}",
                            "Marked as listed from intake.",
                            module="Inventory & Listing Intake", priority="low",
                            related_id=sku)
            st.success(f"Marked {len(to_list)} item(s) as listed and logged tasks.")
            st.rerun()


def _restock_tab() -> None:
    # ── Restock queue: new-arrival items already in the catalogue (from Auto Creation) ──
    st.markdown(section_label("🔁 Restock queue — new arrivals already in your catalogue"),
                unsafe_allow_html=True)
    queue = db.get_restock()
    if not queue:
        st.caption("When Auto Item Creation detects a new-arrival item that already exists in your "
                   "Amazon catalogue, it lands here with an 'add stock' alarm until it's live with "
                   "stock — then it flips to 'done' automatically.")
    else:
        cqa = st.columns([1, 3])
        if cqa[0].button("🔄 Re-check stock", use_container_width=True):
            with st.spinner("Checking Amazon listing status & stock…"):
                for r in queue:
                    try:
                        info = client().confirm_listing(r["sku"])
                        s = str(info.get("status", "")).lower()
                        if any(x in s for x in ("buyable", "active")):
                            db.set_restock_status(r["sku"], "done", info.get("asin", ""))
                    except Exception:
                        pass
            st.rerun()
        cqa[1].caption("'add stock' clears to 'done' once the listing is live (BUYABLE/ACTIVE) with "
                       "stock. Use ✓ to mark done manually, or ✕ to remove.")
        hdr = st.columns([3, 3, 2, 1, 1])
        for c, l in zip(hdr, ("SKU", "Product", "Status", "", "")):
            c.markdown(f"**{l}**")
        for r in queue:
            rc = st.columns([3, 3, 2, 1, 1])
            rc[0].markdown(f"`{r['sku']}`" + (f"  \n{r['asin']}" if r.get("asin") else ""))
            rc[1].markdown(str(r.get("title") or "")[:48] or "—")
            if r.get("status") == "done":
                rc[2].markdown(badge("✅ done", "green"), unsafe_allow_html=True)
            else:
                rc[2].markdown(badge("🔔 add stock", "amber"), unsafe_allow_html=True)
            if rc[3].button("✓", key=f"rsk_done_{r['sku']}", help="Mark done"):
                db.set_restock_status(r["sku"], "done")
                st.rerun()
            if rc[4].button("✕", key=f"rsk_del_{r['sku']}", help="Remove from queue"):
                db.remove_restock(r["sku"])
                st.rerun()
    st.divider()
    st.markdown(section_label("📦 FBA stock cover (catalogued inventory)"), unsafe_allow_html=True)

    inv, src = inventory_source.fetch_inventory()
    if inv.empty:
        st.info("No inventory source available.")
        return
    _, restock = _classify(inv)
    # Join velocity from listings to show urgency.
    from core.api_client import client
    listings = client().get_my_listings()[["sku", "fba_stock", "daily_velocity"]]
    merged = restock.merge(listings, on="sku", how="left").fillna({"fba_stock": 0, "daily_velocity": 0})
    merged["days_left"] = (merged["fba_stock"] / merged["daily_velocity"].replace(0, 0.1)).round(1)
    merged["needs_restock"] = merged["fba_stock"] < merged["daily_velocity"] * 10
    overrides = db.get_channel_overrides()  # FBA/FBM decisions from Hazmat
    merged["channel"] = merged["sku"].map(lambda s: overrides.get(s, "FBA"))

    st.markdown(f"<p style='color:var(--muted)'>{int(merged['needs_restock'].sum())} catalogued "
                f"items below a 10-day stock cover. <span style='color:var(--muted)'>FBM items "
                f"don't need FBA restock.</span></p>", unsafe_allow_html=True)
    styled_table(
        merged[["sku", "title", "brand", "channel", "fba_stock", "daily_velocity", "days_left", "warehouse_qty"]],
        highlight={"row-danger": lambda r: r["fba_stock"] == 0 and r["channel"] == "FBA",
                   "row-warn": lambda r: 0 < r["fba_stock"] < r["daily_velocity"] * 10},
        badge_cols={"channel": {"FBA": ("FBA", "blue"), "FBM": ("FBM", "violet")}})
    export_buttons(merged, "restock_pending")


def _slug(text: str) -> str:
    return re.sub(r"[^A-Za-z0-9]+", "-", str(text)).strip("-")[:24] or "ITEM"


# Words stripped before sending a title to Amazon's product-type search — brand,
# colours and generic noise that confuse the search (we want the noun, not the dressing).
_TYPE_DROP = {
    "green", "lion", "porodo", "lifestyle", "powerology", "lepresso",
    "black", "white", "gray", "grey", "blue", "red", "pink", "rose", "gold",
    "silver", "dark", "light", "beige", "brown", "navy", "purple", "orange",
    "the", "for", "with", "and", "of", "to", "by",
    "premium", "deluxe", "pro", "max", "mini", "ultra", "smart", "new",
    "in", "1", "2", "3", "4", "5", "6", "in-1", "2-in-1", "3-in-1",
}


def _title_keywords(title: str) -> str:
    """Extract the meaningful nouns from a title to send to Amazon's product-type
    search (drops brand, colour, generic adjectives — keeps the core item words)."""
    words = re.findall(r"[A-Za-z0-9]+", str(title or "").lower())
    keep = [w for w in words if w not in _TYPE_DROP and not re.fullmatch(r"\d+(in1|in)?", w)
            and len(w) >= 2]
    return " ".join(keep[:6])


def _amazon_product_type(title: str, valid_types) -> str | None:
    """Ask Amazon's Product Type Definitions search for the best type by keyword
    (caches per keyword in session). Returns a type name or None."""
    kw = _title_keywords(title)
    if not kw:
        return None
    cache = st.session_state.setdefault("aic_pt_search_cache", {})
    if kw in cache:
        results = cache[kw]
    else:
        try:
            results = client().get_product_types(kw) or []
        except Exception:
            results = []
        cache[kw] = results
    valid = set(valid_types) if valid_types else None
    for r in results:
        if not valid or r in valid:
            return r
    return None


def _guess_product_type(title: str) -> str:
    t = (title or "").lower()
    # Right-hand values are REAL Amazon product types (verified to exist in the
    # UAE marketplace) — guessing a non-existent type causes feed error 4000003.
    rules = [# Primary "what is it" categories FIRST so a bundled accessory word in
             # the title (e.g. 'cup holder ... charging cable') can't hijack the type.
             ("CADDY", ["organizer", "organiser", "seat organizer", "seat organiser",
                        "coin organizer", "coin organiser", "caddy", "tidy"]),
             ("CUP_HOLDER", ["cup holder", "cup holder tray", "drinks holder",
                             "drink holder", "beverage holder", "car cup"]),
             ("ELECTRIC_FAN", ["fan", "neck fan", "waist fan", "table fan",
                               "tower fan", "desk fan", "handheld fan", "bladeless"]),
             ("AIR_COOLER", ["air cooler", "cooler", "air cool", "evaporative",
                             "cooling unit", "mist cooler"]),
             ("POWER_BANK", ["power bank", "powerbank", "power core", "magsafe battery"]),
             ("HEADPHONES", ["earbud", "headphone", "airpods", "earphone", "buds", "headset"]),
             ("SPEAKERS", ["speaker", "soundbar", "partybox", "boombox"]),
             ("CHARGING_ADAPTER", ["charger", "charging dock", "gan", "wall adapter",
                                   "car charger", "power adapter"]),
             ("CABLE", ["cable", "usb cable", "charging cable", "type-c cable",
                        "lightning cable", "aux cable", "cord"]),
             ("CELLULAR_PHONE_CASE", ["phone case", "case", "cover", "sleeve", "bumper"]),
             ("SCREEN_PROTECTOR", ["screen protector", "tempered glass", "screen guard"]),
             ("WATCH_BAND", ["watch band", "watch strap", "watch band"]),
             ("CAMERA", ["camera", "webcam", "action cam"]),
             ("WATCH", ["smartwatch", "smart watch", "watch", "fitness band"])]
    for pt, kws in rules:
        # Word-boundary match so short keywords don't hit inside other words
        # (e.g. 'gan' must not match 'orGANizer', 'cam' must not match 'camera' wrongly).
        if any(re.search(r"\b" + re.escape(k) + r"\b", t) for k in kws):
            return pt
    return "GENERIC"


# Price-list columns we auto-map to Amazon attribute names (beyond sku/title/price).
_EXTRA_MAP = {
    "color": ["color", "colour"],
    "model_number": ["model", "model number", "model no", "mpn", "modelnumber"],
    "product_description": ["description", "desc", "long description", "details"],
    "bullet_point": ["bullet", "bullets", "features", "key features", "highlights"],
    "external_product_id": ["barcode", "upc", "ean", "gtin"],
    "watt_hours": ["watt hours", "wh", "watt-hours", "watthours"],
    "wattage": ["wattage", "power", "watt"],
    "material": ["material"],
    "connectivity_technology": ["connectivity", "connection", "connectivity technology"],
    "screen_size": ["screen size", "display size", "screen"],
    "compatible_devices": ["compatible", "compatibility", "compatible devices"],
}


def _inorm(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s).lower())


def _normalize_barcode(raw) -> str:
    """Barcode rule from the price list:
       - if it STARTS WITH 0  -> remove the leading 0
       - otherwise            -> remove the last digit
    Handles values read as numbers (e.g. 6291100.0) by stripping to digits first.
    """
    if raw is None:
        return ""
    s = str(raw).strip()
    if s.endswith(".0"):          # pandas may read the code as a float
        s = s[:-2]
    s = re.sub(r"\D", "", s)        # keep digits only
    if not s:
        return ""
    return s[1:] if s.startswith("0") else s[:-1]


def _parse_battery(feature_text) -> dict:
    """Extract battery energy/mAh/voltage from the Feature spec text so lithium
    fields auto-fill (e.g. 'Battery Capacity: 4000mAh (3.7V / 14.8Wh)')."""
    t = str(feature_text or "")
    out = {}
    # Lookbehind excludes a letter OR digit so a SKU like 'PDLFSTF608WH' isn't
    # misread (neither '608WH' nor the '08WH' substring should match).
    m = re.search(r"(?<![A-Za-z0-9])([\d.]+)\s*wh\b", t, re.I)
    if m:
        out["lithium_energy"] = float(m.group(1))
    mah = re.search(r"(?<![A-Za-z0-9])([\d.]+)\s*mah\b", t, re.I)
    if mah:
        out["mah"] = float(mah.group(1))
    v = re.search(r"(?<![A-Za-z0-9])([\d.]+)\s*v\b", t, re.I)
    if v:
        out["voltage"] = float(v.group(1))
    # Estimate Wh from mAh × V if Wh wasn't stated.
    if "lithium_energy" not in out and "mah" in out:
        out["lithium_energy"] = round(out["mah"] / 1000.0 * out.get("voltage", 3.7), 2)
    if re.search(r"lithium|li-?ion|li-?po|rechargeable|mah|\bwh\b|\bbattery\b|"
                 r"battery capacity|built[- ]?in battery|charging time|working time|"
                 r"playtime|play time|standby", t, re.I):
        out["has_battery"] = True
    return out


def _draft_has_battery(d) -> bool:
    """True if the product has/contains a battery (from any source: pricelist +
    merged packaging text). Drives reusability + the battery field block."""
    if not isinstance(d, dict):
        return False
    ex = d.get("extra", {})
    if ex.get("lithium_energy") or ex.get("mah"):
        return True
    return _parse_battery(d.get("feature", "")).get("has_battery", False)


_KNOWN_MATERIALS = [
    "genuine leather", "pu leather", "vegan leather", "leather", "silicone", "tpu",
    "carbon fiber", "carbon fibre", "stainless steel", "zinc alloy", "aluminum alloy",
    "aluminium", "aluminum", "polycarbonate", "tempered glass", "abs", "pc", "nylon",
    "fabric", "canvas", "rubber", "ceramic", "wood", "bamboo", "glass", "plastic",
    "stainless", "steel",
]


def _parse_material(text) -> str:
    """Extract material composition. First an explicit 'Material: …' label, then a
    fallback scan for common materials mentioned anywhere in the specs."""
    t = str(text or "")
    m = re.search(r"(?:product\s+)?materials?\s*[:\-]\s*([A-Za-z0-9 +/&,().-]{2,45})", t, re.I)
    if m:
        return re.split(r"[\n\r]", m.group(1))[0].strip(" .,-")
    found = []
    for k in _KNOWN_MATERIALS:
        if re.search(r"\b" + re.escape(k) + r"\b", t, re.I):
            found.append(k.lower())
    # Drop any match that is a substring of a more specific one already found
    # (e.g. keep 'genuine leather', drop 'leather'; keep 'aluminum alloy', drop 'aluminum').
    final = [k for k in found if not any(k != o and k in o for o in found)]
    deduped = list(dict.fromkeys(m.title() for m in final))
    return " + ".join(deduped[:3])


_WATT_UNIT = {"w": "watts", "kw": "kilowatts", "mw": "milliwatts"}
_VOLT_UNIT = {"v": "volts", "kv": "kilovolts", "mv": "millivolts"}


def _parse_power(text) -> dict:
    """Highest wattage and voltage WITH their units. Per spec: when several values
    are listed, take the HIGHEST, and always include the unit (W/kW, V/mV…)."""
    t = str(text or "")
    out = {}
    # wattage: capture value + unit (kW/mW/W), excluding 'Wh' (battery energy).
    watts = re.findall(r"(?<![A-Za-z0-9])([\d.]+)\s*(kw|mw|w)\b(?!h)", t, re.I)
    if watts:
        v, u = max(watts, key=lambda x: float(x[0]))
        out["wattage"] = float(v)
        out["wattage_unit"] = _WATT_UNIT.get(u.lower(), "watts")
    volts = re.findall(r"(?<![A-Za-z0-9])([\d.]+)\s*(kv|mv|v)\b", t, re.I)
    if volts:
        v, u = max(volts, key=lambda x: float(x[0]))
        out["voltage"] = float(v)
        out["voltage_unit"] = _VOLT_UNIT.get(u.lower(), "volts")
    return out


# Origins we recognise (map to ISO in sp_api). Only set country from text if known,
# so we never send an unrecognised value that Amazon would reject.
_KNOWN_ORIGINS = {
    "china", "uae", "united arab emirates", "usa", "united states", "india", "japan",
    "germany", "uk", "united kingdom", "south korea", "korea", "taiwan", "vietnam",
    "thailand", "malaysia", "indonesia", "hong kong", "singapore", "france", "italy",
    "spain", "turkey", "saudi arabia",
}


def _parse_carton(text) -> dict:
    """From a 'CARTON DETAILS' cell (e.g. 'Gross Weight (KG): 10.16 Quantity: 14')
    compute the PER-UNIT package & shipping weight = gross weight ÷ units-per-carton.
    The carton DIMENSIONS are for the whole master box (many units), so they are NOT
    used as the individual item's package dimensions."""
    t = str(text or "")
    out = {}
    gw = re.search(r"gross\s*weight[^\d]*([\d.]+)\s*(kg|g)?", t, re.I)
    qty = re.search(r"(?:quantity|qty|units?\s*per\s*carton|pcs?\s*/?\s*carton)[^\d]*([\d.]+)",
                    t, re.I)
    if gw and qty:
        grams_kg = float(gw.group(1))
        if (gw.group(2) or "kg").lower() == "g":
            grams_kg /= 1000.0
        q = float(qty.group(1))
        if q > 0 and grams_kg > 0:
            w = round(grams_kg / q, 3)
            out["package_weight"] = w
            out["shipping_weight"] = w
    return out


def _parse_extra_fields(text) -> dict:
    """Pull additional Amazon fields from the merged pricelist+packaging+manual text:
    country of origin, what's-in-the-box, number of speeds, special features, care
    instructions, safety warnings, and compliance certifications. Stored under the
    exact Amazon attribute names so they fill wherever the product type defines them."""
    t = str(text or "")
    out = {}
    m = re.search(r"(?:made in|country of origin|manufactured in)\s*[:\-]?\s*"
                  r"([A-Za-z][A-Za-z ]{2,28})", t, re.I)
    if m and m.group(1).strip().lower() in _KNOWN_ORIGINS:
        out["country_of_origin"] = m.group(1).strip()
    # What's in the box → included_components
    m = re.search(r"(?:what'?s in the box|package (?:includes?|contents?|list)|"
                  r"in the box|box contains?|package content)\s*[:\-]?\s*(.+)", t, re.I)
    if m:
        line = re.split(r"\n\s*\n", m.group(1))[0]
        parts = [re.sub(r"^\d+\s*[xX]\s*", "", p).strip(" .-")
                 for p in re.split(r"[\n;,••▪]+", line)]
        # Drop manuals / user guides / warranty cards / generic 'device'.
        parts = [p for p in parts if 1 < len(p) <= 40
                 and not any(x in p.lower() for x in _COMP_EXCLUDE)]
        if parts:
            out["included_components"] = ", ".join(dict.fromkeys(parts))[:250]
    # Number of speeds
    m = re.search(r"(\d+)\s*[- ]?\s*(?:speed|speeds|gear|gears|wind\s*speed|fan\s*speed)",
                  t, re.I)
    if m:
        out["number_of_speeds"] = int(m.group(1))
    elif re.search(r"multi[\s-]?speed", t, re.I):
        out["number_of_speeds"] = 3
    # Special features: short, multi-word, descriptive lines (skip specs & manual prose)
    feats = _clean_feature(t)
    _stop = re.compile(
        r"rated|voltage|current|input|charging|time|capacity|material|battery|power|"
        r"weight|dimension|size|model|made in|manual|instruction|warrant|setting|"
        r"content|box|certif|warning|caution|wipe|cloth|wash|immerse|children|compliant",
        re.I)
    # Reject any phrase carrying a measurement/dimension or that is mostly digits
    # (per spec: no raw dimensions/numbers like '170mm' in Special Features).
    _measure = re.compile(r"\d\s*(?:mm|cm|m|in|inch|inches|\"|ml|l|g|kg|w|v|hz|mah|wh)\b", re.I)
    specials = []
    for f in feats:
        if not (4 < len(f) <= 40 and ":" not in f and " " in f):
            continue
        if _stop.search(f) or _measure.search(f):
            continue
        digits = sum(c.isdigit() for c in f)
        if digits and digits / len(f) > 0.25:   # skip number-heavy fragments
            continue
        specials.append(f)
    if specials:
        out["special_feature"] = ", ".join(dict.fromkeys(specials))[:250]
    # Care / maintenance instructions
    m = re.search(r"(?:care instructions?|cleaning(?:\s+instructions?)?|maintenance)\s*"
                  r"[:\-]\s*(.+)", t, re.I)
    if m:
        out["care_instructions"] = re.split(r"\n\s*\n", m.group(1))[0].strip()[:200]
    # Safety warnings / cautions
    m = re.search(r"(?:safety\s+(?:information|warnings?|precautions?)|warnings?|caution)\s*"
                  r"[:\-]\s*(.+)", t, re.I)
    if m:
        out["safety_warning"] = re.split(r"\n\s*\n", m.group(1))[0].strip()[:300]
    # Compliance certifications
    marks = [mk for mk in ["CE", "FCC", "RoHS", "REACH", "UKCA", "RCM", "WEEE", "UL"]
             if re.search(r"\b" + mk + r"\b", t)]
    if marks:
        out["legal_compliance_certification"] = ", ".join(dict.fromkeys(marks))
    return out


# Categories that get NO warranty per the listing spec (cases, screen protectors,
# cables, lanyards/straps and their common synonyms).
_NO_WARRANTY_KW = (
    "phone case", "case", "cover", "sleeve", "bumper", "flip case", "wallet case",
    "screen protector", "tempered glass", "screen guard", "screen film", "protector film",
    "cable", "usb cable", "charging cable", "type-c cable", "lightning cable", "aux cable",
    "cord", "lanyard", "strap", "wrist strap", "neck strap",
)


def _warranty_for(d) -> str:
    """Warranty Description by category: cases/protectors/cables/lanyards → 'No
    Warranty'; everything else → '12 + 12 Months Extended Warranty'."""
    t = ((d.get("title", "") if isinstance(d, dict) else "") + " " +
         (d.get("feature", "") if isinstance(d, dict) else "")).lower()
    if any(k in t for k in _NO_WARRANTY_KW):
        return "No Warranty"
    return db.get_setting("default_warranty_line", "") or "12 + 12 Months Extended Warranty"


def _voltage_for_plug(d) -> str:
    """Accepted Voltage/Frequency by plug: 'no_plug' → blank; otherwise the
    broadest standard that covers UAE type-G mains (220–240V 50Hz)."""
    plug = ""
    if isinstance(d, dict):
        plug = str(d.get("extra", {}).get("power_plug_type", "") or "").lower()
    if plug in ("", "no_plug", "no plug", "none"):
        return ""
    return "100v_240v_50hz_60hz"


def _format_keywords(kw) -> str:
    """Backend Generic Keywords: dedup, semicolon-separated, ≤500 chars."""
    parts = re.split(r"[;,\n]", str(kw or ""))
    seen, low = [], set()
    for p in parts:
        p = p.strip()
        if p and p.lower() not in low:
            low.add(p.lower())
            seen.append(p)
    out = ""
    for p in seen:
        cand = (out + ";" + p) if out else p
        if len(cand) <= 500:
            out = cand
        else:
            break
    return out


def _augment_keywords(d: dict, kw) -> str:
    """Pad the keyword set with relevant product-type × modifier combinations so
    BOTH backend keyword blocks can be filled toward 500 chars each."""
    base, low = [], set()
    for p in re.split(r"[;,\n]", str(kw or "")):
        p = p.strip()
        if p and p.lower() not in low:
            low.add(p.lower())
            base.append(p)
    title = (d.get("title", "") or "").lower()
    head = next((h for h in ["waist fan", "neck fan", "fan", "cooler", "power bank",
                "charger", "cable", "phone case", "case", "speaker", "earbuds",
                "headphones", "watch"] if h in title), "")
    if not head:
        toks = [w for w in re.findall(r"[a-z]{3,}", title)
                if w not in _MN_DROP and w not in ("porodo", "lifestyle")]
        head = toks[-1] if toks else "device"
    mods = ["portable", "rechargeable", "mini", "usb", "cordless", "handheld",
            "personal", "outdoor", "summer", "travel", "wearable", "quiet",
            "lightweight", "desk", "office", "home", "camping", "gym", "kids",
            "compact", "cooling"]
    for m in mods:
        cand = f"{m} {head}".strip()
        if cand and cand.lower() not in low:
            low.add(cand.lower())
            base.append(cand)
    return ";".join(base)


def _split_keywords(kw):
    """Split keywords into TWO ≤500-char semicolon blocks (Generic Keywords field 1
    + the 'Add More' field 2), deduped, no overlap."""
    parts, seen, low = re.split(r"[;,\n]", str(kw or "")), [], set()
    for p in parts:
        p = p.strip()
        if p and p.lower() not in low:
            low.add(p.lower())
            seen.append(p)
    blocks, cur, i = ["", ""], 0, 0
    for p in seen:
        cand = (blocks[cur] + ";" + p) if blocks[cur] else p
        if len(cand) <= 500:
            blocks[cur] = cand
        elif cur == 0:
            cur = 1
            blocks[cur] = p
        else:
            break
    return blocks[0], blocks[1]


def _extra_colmap(df: pd.DataFrame) -> dict:
    """Map Amazon-field -> dataframe column for whatever the price list contains."""
    norm = {_inorm(c): c for c in df.columns}
    out = {}
    for field, aliases in _EXTRA_MAP.items():
        for a in aliases:
            if _inorm(a) in norm:
                out[field] = norm[_inorm(a)]
                break
        else:
            for a in aliases:
                hit = next((orig for nc, orig in norm.items() if _inorm(a) in nc), None)
                if hit:
                    out[field] = hit
                    break
    return out


def _brand_defaults() -> dict:
    return {
        "brand_owner": db.get_setting("brand_owner_mode", "0") == "1",
        "brand": db.get_setting("default_brand", ""),
        "manufacturer": db.get_setting("default_manufacturer", ""),
        "country": db.get_setting("default_country_of_origin", "China"),
    }


# House brands the seller is authorized for (brand owner / distributor). Detected
# from the title so each item lists under its real brand — and so the policy check
# never flags these as third-party trademark issues.
_HOUSE_BRANDS = ["Green Lion", "Porodo", "Powerology", "LePresso", "Porodo Lifestyle"]


def _seller_brands() -> list:
    extra = [b.strip() for b in (db.get_setting("house_brands", "") or "").split(",") if b.strip()]
    dflt = db.get_setting("default_brand", "")
    return list(dict.fromkeys(_HOUSE_BRANDS + ([dflt] if dflt else []) + extra))


def _detect_brand(title: str) -> str:
    """Pick the brand from the title (longest house-brand name that appears)."""
    t = str(title or "")
    for b in sorted(_seller_brands(), key=len, reverse=True):
        if b and re.search(r"\b" + re.escape(b) + r"\b", t, re.I):
            return b
    return ""


def _apply_defaults(draft: dict) -> dict:
    d = dict(draft)
    bd = _brand_defaults()
    if not str(d.get("brand", "")).strip():
        # Brand from the title first (Green Lion / Porodo / …), else the default.
        d["brand"] = _detect_brand(d.get("title", "")) or bd["brand"]
    d.setdefault("extra", {})
    return d


def _relax_for_brand_owner(reqs: list) -> list:
    """Brand owners are GTIN-exempt → barcode fields become optional."""
    if not _brand_defaults()["brand_owner"]:
        return reqs
    out = []
    for f in reqs:
        g = dict(f)
        if g["name"] in ("external_product_id", "external_product_id_type"):
            g["required"] = False
        out.append(g)
    return out


# Amazon attribute names already handled by our curated CORE fields (so we don't
# render them twice when merging with the live schema).
_CORE_COVERED = {
    "item_name", "brand", "manufacturer", "bullet_point", "product_description",
    "color", "color_name", "country_of_origin", "condition_type",
    "external_product_id", "externally_assigned_product_identifier",
    "external_product_id_type", "list_price", "purchasable_offer", "standard_price",
    "fulfillment_availability", "quantity", "main_image_url",
    "main_product_image_locator", "merchant_suggested_asin",
    "parentage_level", "variation_theme", "child_relationship_type",
}


# Smart defaults for category fields Amazon commonly enforces (so they auto-fill).
_SMART_DEFAULTS = {
    "model_number": lambda d: d.get("sku", ""),
    "oem_equivalent_part_number": lambda d: d.get("sku", ""),
    "part_number": lambda d: d.get("sku", ""),
    "manufacturer_part_number": lambda d: d.get("sku", ""),
    "warranty_description": lambda d: _warranty_for(d),
    "seller_warranty_description": lambda d: _warranty_for(d),
    "included_components": lambda d: _default_components(d),
    "supplier_declared_dg_hz_regulation": lambda d: "not_applicable",
    "is_oem_authorized": lambda d: "true",
    "is_oem_sourced_product": lambda d: "false",
    "power_plug_type": lambda d: (d.get("extra", {}).get("power_plug_type") or "no_plug"),
    # Plug logic: no plug → leave voltage/frequency blank; with a plug → match it.
    "accepted_voltage_frequency": lambda d: _voltage_for_plug(d),
    # Hardcoded per spec.
    "item_package_quantity": lambda d: 1,
    "fulfillment_channel": lambda d: "FBA",
    # Reusability: Rechargeable when a battery is present, else blank.
    "reusability": lambda d: "Rechargeable" if _draft_has_battery(d) else "",
    "material": lambda d: _parse_material(d.get("feature", "")),
    "material_type": lambda d: _parse_material(d.get("feature", "")),
    "wattage": lambda d: d.get("extra", {}).get("wattage", "") or "",
    "voltage": lambda d: d.get("extra", {}).get("voltage", "") or "",
    # Presentation / identity fields (spec). Model Name = alphanumeric SKU code
    # (excludes brand/type/colour/size by definition); part number defaults to it.
    "model_name": lambda d: _model_name(d),
    "style": lambda d: _derive_presentation(d)["style"],
    "size": lambda d: _derive_presentation(d)["size"],
    "form_factor": lambda d: _derive_presentation(d)["form_factor"],
    "power_source_type": lambda d: _derive_presentation(d)["power_source_type"],
    "control_method": lambda d: _derive_presentation(d)["control_method"],
    "item_weight": lambda d: _net_weight(d),
    "item_weight_unit": lambda d: "kilograms",
    "shipping_weight_unit": lambda d: "kilograms",
    "dimension_unit": lambda d: "centimeters",
    "package_weight_unit": lambda d: "kilograms",
    "package_dimension_unit": lambda d: "centimeters",
    "number_of_items": lambda d: 1,
    "number_of_boxes": lambda d: 1,
    "required_assembly": lambda d: "No",
    # ELECTRIC_FAN (and some other types) use a boolean field titled "Required
    # Assembly" named is_assembly_required, with enum [False, True] — not "No".
    "is_assembly_required": lambda d: False,
    # Battery yes/no flags follow whether the item actually has a battery (so a
    # non-battery product of the same type answers False, not a wrong True).
    "batteries_required": lambda d: _draft_has_battery(d),
    "batteries_included": lambda d: _draft_has_battery(d),
    "includes_rechargable_battery": lambda d: _draft_has_battery(d),   # Amazon's spelling
    "includes_rechargeable_battery": lambda d: _draft_has_battery(d),
    "lithium_battery_packaging": lambda d: "batteries_contained_in_equipment",
    # Common type-specific fields (audio etc.) — sensible defaults so the listing
    # completes; the user can refine them in the editor.
    "unit_count": lambda d: 1,
    "speaker_type": lambda d: "Portable",
    "speaker_amplification_type": lambda d: "Active",
    "connectivity_technology": lambda d: "Bluetooth",
    "has_multiple_battery_powered_components": lambda d: "false",
    "contains_battery_or_cell": lambda d: "battery",
    "number_of_lithium_ion_cells": lambda d: 1,
    "number_of_batteries": lambda d: 1,
    "battery_cell_composition": lambda d: "lithium_ion",
    "battery_type": lambda d: "nonstandard_battery",
    # Amazon's conditional rule rejects 'installed_in_equipment' for lithium cells;
    # 'installed_in_vessel' is the value that validates + creates the listing.
    "battery_installation_device_type": lambda d: "installed_in_vessel",
    "lithium_packaging": lambda d: "batteries_contained_in_equipment",
    # Energy unit follows what we extracted: mAh capacity (preferred per spec)
    # or watt-hours when only Wh was stated.
    "lithium_energy_unit": lambda d: ("milliampere_hour"
                                      if d.get("extra", {}).get("energy_is_mah")
                                      else "watt_hours"),
    "battery_weight_unit": lambda d: "grams",
    "lithium_weight_unit": lambda d: "grams",
    # Battery physical weight: a minor safe default (30 g, in the 10–50 g range)
    # unless explicitly stated, per spec.
    "battery_weight": lambda d: _est_battery_weight(d),
    "lithium_weight": lambda d: _est_battery_weight(d),
}


# Words stripped when deriving the Model Name (brand, colours, sizes, generic
# product-type/descriptor words) — what remains is the distinctive model name.
_MN_DROP = {
    "porodo", "lifestyle", "powerology", "lepresso", "anker", "baseus",
    "black", "white", "gray", "grey", "blue", "red", "pink", "green", "silver",
    "gold", "dark", "light", "purple", "orange", "beige", "navy", "rose",
    "fan", "cooler", "cooling", "ice", "air", "waist", "neck", "portable", "mini",
    "electric", "handheld", "wearable", "charger", "charging", "cable", "case",
    "cover", "sleeve", "watch", "smartwatch", "smart", "power", "bank", "powerbank",
    "speaker", "headphone", "headphones", "earbud", "earbuds", "earphone", "adapter",
    "hub", "stand", "holder", "mount", "light", "lamp", "kit", "wireless", "bluetooth",
    "rechargeable", "usb", "typec", "type", "with", "for", "and", "the", "desktop",
    "table", "tower", "pedestal", "clip", "strong", "double", "multi", "speed",
    "cup", "suction", "luxe2",
}
_MN_SIZE = re.compile(r"^\d+(?:mah|w|v|wh|gb|tb|mm|cm|inch|ml|l|hz|k)?$", re.I)


# Things that must NEVER appear in Included Components (not box contents / generic).
_COMP_EXCLUDE = ("user manual", "manual", "user guide", "instruction", "instructions",
                 "quick start", "warranty card", "warranty", "device", "catalogue",
                 "catalog", "leaflet")


def _main_item_name(d: dict) -> str:
    """The product's display name with the brand stripped (e.g. 'Chillo Zen Neck Fan')
    — used as the main entry in Included Components."""
    title = str(d.get("title", "") or "")
    brand_toks = {w.lower() for w in (str(d.get("brand", "") or "") + " porodo lifestyle").split()}
    words = [w for w in title.split() if w.lower().strip(",") not in brand_toks]
    return " ".join(words).strip() or title or "Product"


def _default_components(d: dict) -> str:
    """Included Components = the product itself + charging cable. Never 'Device'
    or 'User Manual' (per the listing rules)."""
    return f"{_main_item_name(d)}, USB-C Charging Cable"


def _model_name(d: dict) -> str:
    """The product's specific model name (e.g. 'Chillo Frost') — the distinctive
    words left after removing brand, colour, size and generic product-type words."""
    title = str(d.get("title", "") or "")
    brand_toks = {w.lower() for w in str(d.get("brand", "") or "").split()}
    kept = []
    for w in re.findall(r"[A-Za-z0-9]+", title):
        lw = w.lower()
        if lw in _MN_DROP or lw in brand_toks or _MN_SIZE.match(w):
            continue
        kept.append(w)
    name = " ".join(kept[:3]).strip()
    return name or d.get("sku", "")


def _derive_presentation(d: dict) -> dict:
    """Infer presentation/spec fields from the title + specs: style, size,
    form factor, power source, control method (read like an Amazon tooltip)."""
    t = ((d.get("title", "") if isinstance(d, dict) else "") + " " +
         (d.get("feature", "") if isinstance(d, dict) else "")).lower()
    style = next((w.title() for w in ["minimalist", "modern", "sleek", "classic",
                 "industrial", "premium", "ergonomic", "portable", "compact"] if w in t),
                 "Modern")
    inch = re.search(r"(\d+(?:\.\d+)?)\s*(?:inch|\")", t)
    if inch:
        size = f"{inch.group(1)}-Inch"
    else:
        size = next((w.title() for w in ["compact", "portable", "mini", "large",
                    "medium", "small"] if w in t), "One Size")
    ff_map = [("tower", "Tower"), ("pedestal", "Pedestal"), ("ceiling", "Ceiling"),
              ("wall", "Wall Mounted"), ("table", "Table"), ("desk", "Desktop"),
              ("box fan", "Box"), ("neck", "Neck"), ("waist", "Waist"),
              ("handheld", "Handheld"), ("clip", "Clip-On"), ("wearable", "Wearable")]
    form_factor = next((v for k, v in ff_map if k in t), "Portable")
    if _draft_has_battery(d):
        power_source = "Battery Powered"
    elif "solar" in t:
        power_source = "Solar Powered"
    elif re.search(r"\bplug\b|mains|corded|220v|240v|110v", t):
        power_source = "Corded Electric"
    elif "usb" in t:
        power_source = "USB"
    else:
        power_source = "Battery Powered"
    if "remote" in t:
        control = "remote"
    elif "touch" in t:
        control = "touch"
    elif re.search(r"\bapp\b|application|smartphone control", t):
        control = "application"
    elif "voice" in t:
        control = "voice"
    elif "gesture" in t:
        control = "gesture"
    else:
        control = "push_button"
    return {"style": style, "size": size, "form_factor": form_factor,
            "power_source_type": power_source, "control_method": control}


def _net_weight(d: dict) -> float:
    """Net product weight (kg) ≈ package/shipping weight minus packaging (~15%)."""
    ex = d.get("extra", {}) if isinstance(d, dict) else {}
    pw = ex.get("package_weight") or ex.get("shipping_weight")
    try:
        pw = float(pw)
    except (TypeError, ValueError):
        pw = 0.0
    return round(pw * 0.85, 3) if pw > 0 else ""


def _est_battery_weight(draft) -> float:
    """Battery weight in GRAMS. Per spec: a minor safe default of 30 g (within
    the 10–50 g range) unless the specs explicitly state a weight."""
    ex = draft.get("extra", {}) if isinstance(draft, dict) else {}
    stated = ex.get("battery_weight_g")
    try:
        if stated and float(stated) > 0:
            return round(float(stated), 2)
    except (TypeError, ValueError):
        pass
    return 30.0

# Nested attributes → expand into simple value/unit form fields (mapped back in sp_api).
_NESTED_FIELDS = {
    "website_shipping_weight": [
        {"name": "shipping_weight", "label": "Shipping weight", "type": "number", "required": True},
        {"name": "shipping_weight_unit", "label": "Weight unit", "type": "select",
         "options": ["kilograms", "grams", "pounds", "ounces"], "required": True,
         "default": "kilograms"},
    ],
    "item_depth_width_height": [
        {"name": "item_length", "label": "Item length (depth)", "type": "number", "required": True},
        {"name": "item_width", "label": "Item width", "type": "number", "required": True},
        {"name": "item_height", "label": "Item height", "type": "number", "required": True},
        {"name": "dimension_unit", "label": "Item dimension unit", "type": "select",
         "options": ["centimeters", "inches"], "required": True, "default": "centimeters"},
    ],
    "item_package_weight": [
        {"name": "package_weight", "label": "Package weight", "type": "number", "required": True},
        {"name": "package_weight_unit", "label": "Package weight unit", "type": "select",
         "options": ["kilograms", "grams", "pounds", "ounces"], "required": True,
         "default": "kilograms"},
    ],
    "item_package_dimensions": [
        {"name": "package_length", "label": "Package length", "type": "number", "required": True},
        {"name": "package_width", "label": "Package width", "type": "number", "required": True},
        {"name": "package_height", "label": "Package height", "type": "number", "required": True},
        {"name": "package_dimension_unit", "label": "Package dimension unit", "type": "select",
         "options": ["centimeters", "inches"], "required": True, "default": "centimeters"},
    ],
    "battery": [
        {"name": "battery_cell_composition", "label": "Battery cell composition", "type": "select",
         "options": ["lithium_ion", "lithium_polymer", "lithium_metal", "alkaline",
                     "NiMh", "NiCAD", "lead_acid"], "required": True, "default": "lithium_ion"},
        {"name": "battery_weight", "label": "Battery weight", "type": "number", "required": True},
        {"name": "battery_weight_unit", "label": "Battery weight unit", "type": "select",
         "options": ["grams", "kilograms", "ounces", "pounds"], "required": True,
         "default": "grams"},
    ],
    "lithium_battery": [
        {"name": "lithium_energy", "label": "Lithium energy content (mAh)", "type": "number",
         "required": True},
        {"name": "lithium_energy_unit", "label": "Energy unit", "type": "select",
         "options": ["milliampere_hour", "watt_hours", "kilowatt_hours"], "required": True,
         "default": "milliampere_hour"},
        {"name": "lithium_packaging", "label": "Lithium packaging", "type": "select",
         "options": ["batteries_contained_in_equipment", "batteries_packed_with_equipment",
                     "batteries_only"], "required": True,
         "default": "batteries_contained_in_equipment"},
        {"name": "lithium_weight", "label": "Lithium battery weight", "type": "number",
         "required": True},
        {"name": "lithium_weight_unit", "label": "Lithium weight unit", "type": "select",
         "options": ["grams", "kilograms", "ounces"], "required": True, "default": "grams"},
    ],
    "num_batteries": [
        {"name": "number_of_batteries", "label": "Number of batteries", "type": "number",
         "required": True, "default": 1},
        {"name": "battery_type", "label": "Battery type", "type": "select",
         "options": ["nonstandard_battery", "aa", "aaa", "9v", "lithium_ion"], "required": True,
         "default": "nonstandard_battery"},
    ],
    "battery_installation_device_type": [
        {"name": "battery_installation_device_type", "label": "Battery installation", "type": "select",
         "options": ["installed_in_vessel", "installed_in_equipment", "not_installed",
                     "installed_in_vehicle"], "required": True, "default": "installed_in_vessel"},
    ],
}


def _core_fields() -> list:
    """Our curated set that auto-fills + maps to SP-API correctly (item_name,
    brand, bullets, description, price, qty, barcode, image, colour, channel…)."""
    from core import mock_data
    fields = [dict(f) for f in mock_data.listing_requirements("GENERIC")]
    if not any(f["name"] == "color" for f in fields):
        fields.append({"name": "color", "label": "Color", "required": True, "type": "text"})
    return fields


# Unit fields → a dropdown of sensible units (the user just selects), chosen by
# what the field measures. Used when a *_unit field isn't a schema enum.
def _unit_options_for(name: str):
    n = name.lower()
    if any(w in n for w in ("length", "width", "height", "depth", "dimension", "size")):
        return ["centimeters", "millimeters", "meters", "inches", "feet"], "centimeters"
    if "weight" in n:
        return ["kilograms", "grams", "milligrams", "pounds", "ounces"], "kilograms"
    if any(w in n for w in ("volume", "capacity", "liquid")):
        return ["liters", "milliliters", "fluid_ounces", "gallons", "cubic_centimeters"], "liters"
    if "voltage" in n or "volt" in n:
        return ["volts", "kilovolts", "millivolts"], "volts"
    if "watt" in n or "power" in n:
        return ["watts", "kilowatts", "milliwatts"], "watts"
    if any(w in n for w in ("energy", "mah", "ampere")):
        return ["milliampere_hour", "watt_hours", "kilowatt_hours"], "milliampere_hour"
    if "frequency" in n:
        return ["hertz", "kilohertz", "megahertz", "gigahertz"], "hertz"
    return None, None


def _expand_field(name: str, full_by_name: dict) -> list:
    """Turn an Amazon attribute name into renderable form field(s)."""
    if name in _NESTED_FIELDS:
        return [dict(f) for f in _NESTED_FIELDS[name]]
    f = full_by_name.get(name)
    if f:
        return [dict(f)]
    # Not in schema (e.g. a unit name pulled from a reject message). If it's a UNIT
    # field, render a DROPDOWN of sensible units instead of a blank text box.
    if "unit" in name.lower():
        opts, default = _unit_options_for(name)
        if opts:
            return [{"name": name, "label": name.replace("_", " ").title(), "type": "select",
                     "options": opts, "default": default, "required": True}]
    return [{"name": name, "label": name.replace("_", " ").title(),
             "required": True, "type": "text"}]


# Fields Amazon enforces at feed time but DOESN'T mark required in the schema.
# We pre-include any that the product type actually defines, so the first push is complete.
_COMMON_ENFORCED = [
    "model_number", "oem_equivalent_part_number",
    "power_plug_type", "included_components", "warranty_description",
    "required_assembly", "is_assembly_required",
    "number_of_items", "number_of_boxes", "is_oem_authorized",
    "supplier_declared_dg_hz_regulation", "item_depth_width_height",
    "item_package_dimensions", "item_package_weight", "website_shipping_weight",
    # Battery yes/no flags: ALWAYS included when the type defines them (a type that
    # requires them needs an answer even if our text-detection didn't see a battery).
    # NOTE: Amazon's real attribute is mis-spelled 'rechargable' (no 'e'); we list
    # both spellings — only the one the product type defines is actually added.
    "includes_rechargable_battery", "includes_rechargeable_battery",
    "batteries_required", "batteries_included",
]
_BATTERY_ENFORCED = [
    "battery", "num_batteries", "lithium_battery", "contains_battery_or_cell",
    "has_multiple_battery_powered_components", "number_of_lithium_ion_cells",
    "battery_installation_device_type", "lithium_battery_packaging",
]
# Optional-but-valuable fields: filled automatically whenever the product type
# defines them (material, reusability, power specs, backend keywords, etc.) so the
# listing is as complete as possible — never left blank when data exists.
_OPTIONAL_ENRICH = [
    "material", "reusability", "wattage", "voltage", "generic_keyword",
    "item_package_quantity", "number_of_items", "color",
    # Extra fields pulled from packaging / user manual when the type defines them.
    "included_components", "number_of_speeds", "special_feature", "country_of_origin",
    "care_instructions", "safety_warning", "legal_compliance_certification",
    # Presentation / identity / spec fields.
    "model_name", "part_number", "style", "size", "form_factor", "power_source_type",
    "control_method", "item_weight", "item_shape", "mounting_type",
    # Audio / common type-specific fields (filled when the type defines them).
    "unit_count", "speaker_type", "speaker_amplification_type", "connectivity_technology",
    # NOTE: 'Item Highlight' (title_differentiation) is NOT sent — Amazon rejects it
    # via the API as "currently unsupported"; it's a Seller-Central-only field. We
    # still generate the text and show it in the editor for manual entry.
]

# Map an Amazon error-message field TITLE (lowercased) → the attribute name, so a
# rejection like "'Includes Rechargeable Battery' is required" auto-adds the field.
_TITLE_TO_ATTR = {
    "includes rechargeable battery": "includes_rechargable_battery",
    "rechargeable battery included": "includes_rechargable_battery",
    "includes rechargable battery": "includes_rechargable_battery",
    "are batteries required": "batteries_required",
    "are batteries included": "batteries_included",
    "battery cell composition": "battery_cell_composition",
    "number of lithium ion cells": "number_of_lithium_ion_cells",
    "number of lithium-ion cells": "number_of_lithium_ion_cells",
    "lithium battery packaging": "lithium_battery_packaging",
    "battery type": "battery_type",
    "number of batteries": "num_batteries",
    "battery weight": "battery_weight",
    "required assembly": "is_assembly_required",
    "supplier declared dg hz regulation": "supplier_declared_dg_hz_regulation",
    "country of origin": "country_of_origin",
    "warranty description": "warranty_description",
    "manufacturer": "manufacturer",
    "model name": "model_name",
    "model number": "model_number",
    "part number": "part_number",
    "power plug": "power_plug_type",
    "number of items": "number_of_items",
    "number of boxes": "number_of_boxes",
    "speaker amplification": "speaker_amplification_type",
    "speaker type": "speaker_type",
    "unit count": "unit_count",
    # Item-dimension unit → pull in the WHOLE nested item-dimensions block
    # (length/width/height + unit), since Amazon needs the values too.
    "item length unit": "item_depth_width_height",
    "item width unit": "item_depth_width_height",
    "item height unit": "item_depth_width_height",
    "item dimensions unit": "item_depth_width_height",
    "item length": "item_depth_width_height",
    "item width": "item_depth_width_height",
    "item height": "item_depth_width_height",
    # Package-dimension unit/value → pull in the WHOLE package-dimensions block
    # (length/width/height + unit), not a lone dead 'Package Height Unit' field.
    "package length unit": "item_package_dimensions",
    "package width unit": "item_package_dimensions",
    "package height unit": "item_package_dimensions",
    "package dimensions unit": "item_package_dimensions",
    "package length": "item_package_dimensions",
    "package width": "item_package_dimensions",
    "package height": "item_package_dimensions",
    # Automotive category required field.
    "automotive fit type": "automotive_fit_type",
}


def _attrs_from_messages(issues) -> set:
    """Extract attribute names from rejection/validation messages by reading the
    quoted field title. Known titles (incl. Amazon mis-spellings) map via the table;
    anything else falls back to the snake_case of the title, which is the attribute
    name for most fields (e.g. 'Speaker Type' -> speaker_type). The discovery step
    then keeps only names the product type actually defines."""
    out = set()
    for i in issues or []:
        msg = i.get("message", "") if isinstance(i, dict) else str(i)
        # ignore count/format errors — only 'X is required but missing' adds fields
        for title in re.findall(r"'([^']+)'\s+is required", msg, re.I):
            t = title.strip().lower()
            if t in _TITLE_TO_ATTR:
                out.add(_TITLE_TO_ATTR[t])
            else:
                snake = re.sub(r"[^a-z0-9]+", "_", t).strip("_")
                if 2 <= len(snake) <= 45 and not snake.isdigit():
                    out.add(snake)
    return out


def _hybrid_requirements(pt: str, has_battery: bool = False) -> list:
    """LIVE form = curated CORE + Amazon's required fields + commonly-enforced fields
    (that this product type defines) + battery block (if the item has a battery) +
    anything discovered from a previous validation. Nested fields expand to simple
    value/unit inputs; smart defaults pre-fill most so the first push is complete."""
    core = _core_fields()
    covered = set(_CORE_COVERED) | {f["name"] for f in core}
    try:
        live = client().get_listing_requirements(pt)
    except Exception:
        return core
    full_by_name = {f["name"]: f for f in live}
    discovered = set(st.session_state.get("aic_more", {}).get(pt, set()))
    # Names that really refer to the nested item-dimensions block → expand the whole
    # block (length/width/height + unit), not a lone dead 'item_length_unit' field.
    _dim_alias = {"item_length_unit", "item_width_unit", "item_height_unit",
                  "item_dimensions_unit", "item_length", "item_width", "item_height"}
    if discovered & _dim_alias:
        discovered = (discovered - _dim_alias) | {"item_depth_width_height"}
    # Same for package-dimension fragments (e.g. a self-healed 'package_height_unit')
    # → expand the whole package-dimensions block instead of a lone dead unit field.
    _pkg_dim_alias = {"package_length_unit", "package_width_unit", "package_height_unit",
                      "package_dimensions_unit", "package_length", "package_width",
                      "package_height", "package_size"}
    if discovered & _pkg_dim_alias:
        discovered = (discovered - _pkg_dim_alias) | {"item_package_dimensions"}
    need = [f["name"] for f in live if f.get("required")] + list(discovered)
    enforce = list(_COMMON_ENFORCED) + (list(_BATTERY_ENFORCED) if has_battery else [])
    need += [nm for nm in enforce if nm in full_by_name]
    # Automotive types feed-enforce 'automotive_fit_type' even though the schema
    # marks it optional — include it so the first push isn't rejected for it.
    auto_enf = [n for n in ("automotive_fit_type",) if n in full_by_name]
    need += auto_enf
    # Optional enrich fields: include when the type defines them (auto-filled,
    # not required, so they never block a push).
    need += [nm for nm in _OPTIONAL_ENRICH if nm in full_by_name]

    # Fields Amazon explicitly rejected for (discovered) + automotive-enforced are
    # marked REQUIRED in our form so the readiness check catches them (an item is
    # never shown 'ready' while one of these is still blank).
    force_req = set(discovered) | set(auto_enf)
    extra, seen = [], set()
    for name in need:
        if name in covered or name in seen:
            continue
        seen.add(name)
        for fld in _expand_field(name, full_by_name):
            if fld["name"] in covered or fld["name"] in {x["name"] for x in extra}:
                continue
            if name in force_req or fld["name"] in force_req:
                fld = dict(fld, required=True)
                if fld.get("type") == "select" and not fld.get("default"):
                    fld["default"] = (fld.get("options") or [""])[0]
            extra.append(fld)
    return core + extra


def _coerce_field_value(field: dict, val):
    """Coerce a prefilled value to the schema's value type so it matches the
    form widget's options and Amazon's expected JSON type.

    - boolean fields (enum [False, True]) → real bool (string 'true'/'yes'/'1'
      → True), so the selectbox prefills correctly instead of dropping to False.
    - integer fields → int (Amazon rejects 1.0 for an integer attribute).
    """
    is_bool = field.get("boolean") or (
        field.get("options") and all(isinstance(o, bool) for o in field["options"]))
    if is_bool:
        if isinstance(val, bool):
            return val
        return str(val).strip().lower() in ("true", "yes", "1", "y", "required")
    if field.get("integer"):
        try:
            return int(float(val))
        except (TypeError, ValueError):
            return val
    return val


def _num(x, default: float = 0.0) -> float:
    """Best-effort number from any value, WITHOUT crashing. Accepts a clean numeric
    string (optionally with 'AED'/'$'/commas); anything messy (a spec blob like
    'Dimensions (CM): 36.5*30*35\\nGross Weight…', empty, NaN) → `default`. We do NOT
    pluck a stray number out of a sentence — a 0 price is flagged as missing so the
    user fixes it, which is safer than guessing a dimension as the price."""
    if x is None or isinstance(x, bool):
        return default
    if isinstance(x, (int, float)):
        try:
            return default if pd.isna(x) else float(x)
        except (TypeError, ValueError):
            return default
    s = str(x).strip().replace(",", "")
    for tok in ("AED", "aed", "Aed", "$", "USD", "usd"):
        s = s.replace(tok, "")
    s = s.strip()
    try:
        return float(s)
    except (TypeError, ValueError):
        return default


def _prefill(field: dict, draft: dict, opt: dict):
    """Pre-fill a field from sources, then coerce to the schema's value type."""
    return _coerce_field_value(field, _prefill_value(field, draft, opt))


def _prefill_value(field: dict, draft: dict, opt: dict):
    """Pre-fill a field from price-list extras, fetched draft, optimized copy, defaults."""
    name = field["name"]
    extra = draft.get("extra", {})
    if name in extra and str(extra[name]).strip():
        return extra[name]
    bd = _brand_defaults()
    if name == "manufacturer":
        return draft.get("brand") or bd["manufacturer"] or bd["brand"]
    if name == "country_of_origin":
        return bd["country"] or field.get("default", "")
    # Backend search terms: show the generated, semicolon-joined keywords (≤500)
    # right in the form so the field isn't blank.
    if name in ("generic_keyword", "generic_keywords"):
        return _split_keywords(_augment_keywords(draft, opt.get("keywords", "")))[0]
    # Item Highlight (title_differentiation): ONE value, ≤125 chars — pack as many
    # short benefit phrases as fit, comma-separated (Amazon allows a single entry).
    if name == "title_differentiation":
        s = ""
        for h in opt.get("highlights", []):
            cand = (s + ", " + h) if s else h
            if len(cand) <= 125:
                s = cand
            else:
                break
        return s
    src = field.get("source")
    if src == "title":
        return opt.get("title", draft.get("title", ""))
    if src == "bullets":
        return "\n".join(opt.get("bullets", []))
    if src == "description":
        return opt.get("description", "")
    if src == "brand":
        return draft.get("brand", "")
    if src == "price":
        return _num(draft.get("price"))
    if src == "image":
        imgs = draft.get("images") or []
        return imgs[0] if imgs else ""
    if name in _SMART_DEFAULTS:
        return _SMART_DEFAULTS[name](draft)
    return field.get("default", "")


# ---------------------------------------------------------------------------
# Smart price-list parser (handles formatted sheets: header not on row 1,
# section/category rows, multi-price columns, hyperlinked media/image cells).
# ---------------------------------------------------------------------------
_PL_ALIASES = {
    "sku": ["sku", "seller sku", "item code", "code", "model"],
    "title": ["product name", "title", "name", "product", "item name"],
    "feature": ["feature", "features", "specification", "specs", "details", "description"],
    "color": ["color", "colour"],
    "barcode": ["barcode", "ean", "upc", "gtin"],
    "price_rrp": ["rrp", "rrpaed", "recommended"],
    "price_mrp": ["mrp", "mrpaed"],
    "price_base": ["price", "priceaed", "unit price", "selling price"],
    "media": ["media", "media link", "medialink"],
    "image": ["image", "mockup", "photo", "picture"],
    "carton": ["carton", "carton details", "packaging"],
}


def _pnorm(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    return re.sub(r"[^a-z0-9]", "", str(x).lower())


def _detect_header_and_map(raw: pd.DataFrame) -> tuple:
    """Find the header row and map logical fields → column index."""
    best = (None, -1, {})
    for ri in range(min(len(raw), 15)):
        colmap = {}
        for ci in range(raw.shape[1]):
            cell = _pnorm(raw.iat[ri, ci])
            if not cell:
                continue
            for field, aliases in _PL_ALIASES.items():
                if field in colmap:
                    continue
                if any(cell == _pnorm(a) or _pnorm(a) in cell for a in aliases):
                    colmap[field] = ci
        if len(colmap) > best[1]:
            best = (ri, len(colmap), colmap)
    hidx, score, colmap = best
    return (hidx if score >= 3 else None), colmap


def _clean_feature(s) -> list:
    """Split a Feature cell into clean lines (fix mangled bullet/encoding chars)."""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return []
    txt = str(s)
    for junk in ["�", "•", "▪", "●", "·", "*"]:
        txt = txt.replace(junk, "\n")
    lines = re.split(r"[\n\r]+", txt)
    out = []
    for ln in lines:
        ln = re.sub(r"^[^A-Za-z0-9]+", "", ln).strip()
        if len(ln) > 1:
            out.append(ln)
    return out


def _parse_pricelist_bytes(data: bytes, is_csv: bool) -> pd.DataFrame:
    """Return a normalized DataFrame: sku,title,feature,color,price,barcode,media,image."""
    if is_csv:
        raw = pd.read_csv(io.BytesIO(data), header=None, dtype=str)
        ws = None
    else:
        raw = pd.read_excel(io.BytesIO(data), header=None, dtype=str)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(data))
            ws = wb[wb.sheetnames[0]]
        except Exception:
            ws = None

    hidx, colmap = _detect_header_and_map(raw)
    if hidx is None:
        # No header row found (e.g. a single-product export with data on row 0).
        # The price lists follow a fixed column template, so map by POSITION and
        # read every row from the top (hidx = -1 → data loop starts at row 0).
        ncols = raw.shape[1]
        positional = {"title": 0, "feature": 1, "sku": 2, "barcode": 3, "image": 4,
                      "color": 6, "price_base": 7, "price_mrp": 8, "price_rrp": 9,
                      "carton": 10, "media": 11}
        colmap = {k: v for k, v in positional.items() if v < ncols}
        hidx = -1
    def cell_at(ri, ci):
        if ci is None:
            return ""
        v = raw.iat[ri, ci]
        return "" if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v).strip()

    def cell(ri, key):
        return cell_at(ri, colmap.get(key))

    # --- PRICE column selection, VALUE-VALIDATED -------------------------------
    # Pick the price column by the user's preference (RRP → MRP → base), but ONLY if
    # it actually holds numbers. Header/positional guesses can land on a specs/carton
    # column (e.g. 'Dimensions… Gross Weight… Quantity'), so if the preferred price
    # columns aren't numeric we find the real numeric price column from the data.
    _data_rows = list(range(hidx + 1, len(raw)))

    def _num_ratio(ci):
        if ci is None:
            return 0.0
        vals = [cell_at(ri, ci) for ri in _data_rows]
        vals = [v for v in vals if v]
        return (sum(1 for v in vals if _num(v, 0) > 0) / len(vals)) if vals else 0.0

    pref_price = [colmap[k] for k in ("price_rrp", "price_mrp", "price_base")
                  if colmap.get(k) is not None]
    price_cols = [ci for ci in pref_price if _num_ratio(ci) >= 0.5]
    if not price_cols:
        # No usable mapped price column → find numeric columns that aren't an
        # id/text/specs column, and take the RIGHT-MOST (RRP sits toward the right
        # in these price lists). This recovers the price even when the header row
        # wasn't detected and the positional guess pointed at the wrong column.
        non_price = {colmap.get(k) for k in ("sku", "barcode", "media", "image",
                     "feature", "title", "color", "carton")
                     if colmap.get(k) is not None}
        numeric = [ci for ci in range(raw.shape[1])
                   if ci not in non_price and _num_ratio(ci) >= 0.6]
        if numeric:
            price_cols = [max(numeric)]

    def _row_price(ri):
        for pc in price_cols:
            v = cell_at(ri, pc)
            if _num(v, 0) > 0:
                return v
        return ""

    def hyperlink(ri, key):
        ci = colmap.get(key)
        if ci is None or ws is None:
            return ""
        try:
            c = ws.cell(row=ri + 1, column=ci + 1)
            if c.hyperlink and c.hyperlink.target:
                return c.hyperlink.target
        except Exception:
            pass
        return ""

    rows = []
    for ri in range(hidx + 1, len(raw)):
        sku = cell(ri, "sku")
        barcode = cell(ri, "barcode")
        title = cell(ri, "title")
        if not sku and not barcode:
            continue  # skip section/category/blank rows
        media = hyperlink(ri, "media") or cell(ri, "media")
        image = hyperlink(ri, "image") or cell(ri, "image")
        if media and not media.startswith("http"):
            media = ""  # text like "Media Link" with no real hyperlink
        if image and not image.startswith("http"):
            image = ""
        rows.append({
            "sku": sku, "title": title, "feature": cell(ri, "feature"),
            "color": cell(ri, "color"), "price": _row_price(ri),
            "barcode": barcode, "media": media, "image": image,
            "carton": cell(ri, "carton"), "_row": ri,
        })
    # Variant rows (blank title/feature/price) inherit from the parent row above,
    # and share a group_id with it (→ a colour variation family).
    # Children inherit the parent's shared specs (carton too → same package weight).
    last = {"title": "", "feature": "", "price": "", "carton": ""}
    gid = -1
    for r in rows:
        own = bool(str(r.get("title") or "").strip())
        if own:
            gid += 1
        r["is_child"] = not own
        r["group_id"] = max(gid, 0)
        for k in ("title", "feature", "price", "carton"):
            if str(r.get(k) or "").strip():
                last[k] = r[k]
            elif last[k]:
                r[k] = last[k]
    return pd.DataFrame(rows)


def _load_pricelist(up, use_sample: bool) -> pd.DataFrame:
    if up is not None:
        return _parse_pricelist_bytes(up.getvalue(), up.name.lower().endswith(".csv"))
    if use_sample:
        from core import mock_data
        s = mock_data.sample_price_list().copy()
        s["feature"] = ""
        s["color"] = ""
        s["barcode"] = ""
        s["image"] = ""
        s = s.rename(columns={"media_link": "media"})
        return s[["sku", "title", "price", "feature", "color", "barcode", "media", "image"]]
    return pd.DataFrame()


def _pdf_text(data: bytes) -> str:
    try:
        from pypdf import PdfReader
        return "\n".join((p.extract_text() or "")
                         for p in PdfReader(io.BytesIO(data)).pages)
    except Exception:
        return ""


def _load_packaging(files) -> list:
    """Packaging files (PDF / CSV / Excel) → list of (label, text). Matched to
    products later by scanning the text for each SKU/barcode."""
    if not files:
        return []
    if not isinstance(files, list):
        files = [files]
    out = []
    for f in files:
        name = f.name.lower()
        try:
            if name.endswith(".pdf"):
                txt = _pdf_text(f.getvalue())
                if txt.strip():
                    out.append((f.name, txt))
            elif name.endswith(".txt"):
                txt = f.getvalue().decode("utf-8", "replace")
                if txt.strip():
                    out.append((f.name, txt))
            else:
                df = (pd.read_csv(f, dtype=str) if name.endswith(".csv")
                      else pd.read_excel(f, dtype=str))
                # one text block per row, prefixed with its sku/barcode for matching
                for _, r in df.iterrows():
                    block = " | ".join(f"{c}: {r[c]}" for c in df.columns
                                       if pd.notna(r[c]) and str(r[c]).strip())
                    if block:
                        out.append((f.name, block))
        except Exception:
            continue
    return out


def _packaging_for(sku: str, barcode: str, packaging: list) -> str:
    """Find the packaging file matching this SKU/barcode (by filename or content)."""
    for name, text in packaging:
        hay_name = str(name)
        if (sku and (sku in text or sku in hay_name)) or \
           (barcode and (barcode in text or barcode in hay_name)):
            return text
    # If exactly one packaging file was given, assume it's for the item.
    if len(packaging) == 1:
        return packaging[0][1]
    return ""


# When only the PACKAGE (box) size is known, estimate the bare ITEM as this
# fraction of the box on each axis — a product always fits inside its package.
# Mirrors _net_weight()'s "item ≈ less than package" rule for weight. The inverse
# (package = item / factor) estimates the box when only the product size is given.
_ITEM_FROM_PACKAGE_FACTOR = 0.9


def _find_dims(text: str, require_unit: bool = False):
    """First length×width×height (+ optional unit) in `text` → (l, w, h, unit) in
    centimeters/inches, or None. Accepts any separator (x * × or the � that PDFs
    sometimes decode the × to) and converts mm → cm. When `require_unit` is True a
    trailing length unit is mandatory (so dates / model numbers aren't misread)."""
    unit = r"(mm|cm|millimet\w*|centimet\w*|inch\w*|in|\")"
    tail = unit + (r"\b" if require_unit else r"?")
    m = re.search(r"([\d.]+)\s*[^\d.\s]{1,3}\s*([\d.]+)\s*[^\d.\s]{1,3}\s*([\d.]+)\s*" + tail,
                  text, re.I)
    if not m:
        return None
    l, w, h = float(m.group(1)), float(m.group(2)), float(m.group(3))
    u = (m.group(4) or "cm").lower()
    if u.startswith("mm") or u.startswith("millim"):
        return round(l / 10, 2), round(w / 10, 2), round(h / 10, 2), "centimeters"
    if u.startswith("in") or u == '"':
        return l, w, h, "inches"
    return l, w, h, "centimeters"


def _labeled_dims(text: str, *keywords):
    """Dimensions on a line introduced by any `keywords` label (e.g. 'Package Size:
    16x12x7 cm'). Looks only at the 40 chars after the label so it can't bleed onto
    the next line. Unit is optional here — the label supplies the context."""
    pat = r"(?:%s)\s*(?:\([^)]*\))?\s*[-:]?\s*([^\n]{0,40})" % "|".join(keywords)
    m = re.search(pat, text, re.I)
    return _find_dims(m.group(1)) if m else None


def _parse_weight(text: str, label: str):
    """Weight following `label` (regex) → (value, unit∈{grams,kilograms}) or None.
    Reads the unit from a trailing token ('382 g'), else a parenthetical in the
    label ('Gross Weight (KG): 9.6'), else defaults to grams."""
    m = re.search(label + r"\s*(\([^)]*\))?[^\d-]*[-:]?\s*([\d.]+)\s*"
                  r"(kg|kgs|kilograms?|g|gram[s]?)?", text, re.I)
    if not m:
        return None
    u = (m.group(3) or m.group(1) or "g").lower()
    return float(m.group(2)), ("kilograms" if ("kg" in u or "kilo" in u) else "grams")


def _parse_dimensions(text: str) -> dict:
    """Pull PACKAGE + ITEM dimensions and weights out of packaging/spec text.

    Package = the box, read straight from the data: an explicit 'package / box /
    carton size' line if present, otherwise the first measured size in the sheet
    (these sheets describe the box). Item = an explicit 'product / item size' line
    if present; otherwise estimated as slightly LESS than the package, since the
    product fits inside its box. If only the product size is given, the package is
    estimated slightly LARGER. Item is always clamped to ≤ package on every axis.

    Weights: 'gross weight' → package/shipping weight; 'net/item/product weight' →
    item weight; a bare 'weight' (no net/gross label) → package/shipping weight.
    Returns flat keys consumed by the form + sp_api.to_sp_attributes()."""
    out = {}
    t = str(text or "")

    # --- dimensions -------------------------------------------------------
    pkg = _labeled_dims(t, r"package\s*size", r"packing\s*size", r"package\s*dimensions?",
                        r"box\s*size", r"carton\s*size", r"gift\s*box", r"outer\s*box")
    item = _labeled_dims(t, r"product\s*size", r"item\s*size",
                         r"product\s*dimensions?", r"item\s*dimensions?", r"unit\s*size")
    if not pkg and not item:
        pkg = _find_dims(t, require_unit=True)   # unlabeled sheet → treat as the box

    f = _ITEM_FROM_PACKAGE_FACTOR

    def _scale(dims, factor):
        l, w, h, u = dims
        return (round(l * factor, 1) or l, round(w * factor, 1) or w,
                round(h * factor, 1) or h, u)

    if pkg and not item:
        item = _scale(pkg, f)            # product is smaller than its box
    elif item and not pkg:
        pkg = _scale(item, 1 / f)        # box is larger than the product

    if pkg:
        out.update({"package_length": pkg[0], "package_width": pkg[1],
                    "package_height": pkg[2], "package_dimension_unit": pkg[3]})
    if item:
        il, iw, ih, iu = item
        if pkg:   # never let the product exceed its package on any axis
            il, iw, ih = min(il, pkg[0]), min(iw, pkg[1]), min(ih, pkg[2])
        out.update({"item_length": il, "item_width": iw, "item_height": ih,
                    "dimension_unit": iu})

    # --- weights ----------------------------------------------------------
    gross = _parse_weight(t, r"gross\s*weight")
    net = _parse_weight(t, r"(?:net|item|product)\s*weight")
    pkg_w = gross or (_parse_weight(t, r"weight") if not net else None)
    if pkg_w:
        out.update({"package_weight": pkg_w[0], "package_weight_unit": pkg_w[1],
                    "shipping_weight": pkg_w[0], "shipping_weight_unit": pkg_w[1]})
    if net:
        out.update({"item_weight": net[0], "item_weight_unit": net[1]})
        if not pkg_w:   # only the product weight is known → box is a bit heavier
            kg = net[0] if net[1] == "kilograms" else net[0] / 1000.0
            gw = round(kg / 0.85, 3)
            out.update({"package_weight": gw, "package_weight_unit": "kilograms",
                        "shipping_weight": gw, "shipping_weight_unit": "kilograms"})
    return out


# Default warranty/support line for the mandatory "What You Get" last bullet.
_DEFAULT_WARRANTY = ("24-Hour Customer Service, Lifetime Technical Support and "
                     "12 + 12 Months Warranty")


def _fallback_highlights(draft: dict) -> list:
    """Short benefit-driven 'Item Highlight' phrases (≤125 chars, not sentences,
    not repeating the item name) derived from the product's features/specs."""
    title = (draft.get("title", "") or "").lower()
    feat = draft.get("feature", "")
    out = []
    sf = _parse_extra_fields(feat).get("special_feature", "")
    for p in sf.split(","):
        p = p.strip()
        if p and p.lower() not in title:
            out.append(p)
    bat = _parse_battery(feat)
    if bat.get("mah"):
        out.append(f"{int(bat['mah'])}mAh Rechargeable Battery")
    m = re.search(r"(\d+)\s*hour", feat, re.I)
    if m:
        out.append(f"Up to {m.group(1)} Hours Runtime")
    mat = _parse_material(feat)
    if mat:
        out.append(f"Durable {mat} Build")
    seen = []
    for p in out:
        p = p[:125].strip()
        if p and p.lower() not in {s.lower() for s in seen}:
            seen.append(p)
    return seen[:6]


def _listing_copy(draft: dict) -> dict:
    """Generate Amazon copy to the seller's standards:
       • Title  : < 200 chars, brand + type + keyword-rich feature phrases (SEO).
       • Bullets: 5-6, each = Bold Lead-in: 2-3 sentences; LAST = 'What You Get:'.
       • Desc   : keyword-dense paragraph.
    Uses Claude when an Anthropic key is set, else a structured rule-based builder."""
    bd = _brand_defaults()
    brand = draft.get("brand") or bd["brand"] or "Porodo"
    name = draft.get("title", "")
    color = draft.get("extra", {}).get("color", "")
    feats = _clean_feature(draft.get("feature", ""))
    warranty = db.get_setting("default_warranty_line", "") or _DEFAULT_WARRANTY
    # Strip any "Free" qualifier before the warranty term (per listing rules).
    warranty = re.sub(r"\bfree\s+", "", warranty, flags=re.I).strip()

    key = db.get_setting("anthropic_api_key", "")
    if key and (name or feats):
        try:
            from core import assistant
            system = ("You are an expert Amazon.ae SEO listing copywriter. You write "
                      "keyword-rich, policy-compliant copy that ranks high and reads naturally. "
                      "Return ONLY valid JSON.")
            prompt = (
                f"Brand: {brand}\nProduct: {name}\nColor: {color}\n"
                f"Warranty/Support line (use verbatim in the last bullet): {warranty}\n"
                f"Product features/specs:\n" + "\n".join("- " + f for f in feats) +
                "\n\nWrite an Amazon.ae listing as JSON with keys title, bullets, highlights, description, keywords.\n\n"
                "TITLE (string): ONE line, MUST be UNDER 200 characters (aim 185-198). Begin with the "
                "Brand, then the product type, then the most important features/specs as "
                "comma/ampersand-separated keyword phrases — mirror how top-selling Amazon listings "
                "pack searchable keywords buyers actually type, using ONLY terms this product truly has. "
                "Title Case. NO promotional words (best, cheap, sale, #1, free shipping, guarantee).\n\n"
                "BULLETS (array of 5 to 6 strings): order by feature importance. Each non-final bullet = "
                "a Bold Lead-in Phrase, then ': ', then 2-3 complete sentences of benefit-led, simple, "
                "readable copy that still weaves in the important keywords (e.g. "
                "'Powerful 2950W Dual Boiler: <2-3 sentences>'). The FINAL bullet MUST begin with "
                f"'What You Get:' and list the product plus this exact text: '{warranty}'.\n\n"
                "HIGHLIGHTS (array of 5 to 6 strings): short benefit-driven feature phrases, NOT full "
                "sentences, each UNDER 100 characters (e.g. 'Hands-Free Waist Cooling', "
                "'Up to 5 Hours Runtime'). Do NOT repeat the brand or the product type already in the "
                "title. Title Case, no ending punctuation.\n\n"
                "DESCRIPTION (string): a keyword-dense paragraph of 4-6 sentences that repeats and varies "
                "the most important search keywords, specs and use-cases (it aids search even if unread).\n\n"
                "KEYWORDS (string): 35-50 comma-separated backend search terms — include synonyms, "
                "use-cases, audiences, materials and the localized variants buyers actually type; "
                "do not repeat the brand. Maximize coverage.\n\n"
                "Return JSON only.")
            txt, status = assistant.complete(system, prompt, max_tokens=1800)
            if status == "ok" and txt and "{" in txt:
                data = json.loads(txt[txt.find("{"):txt.rfind("}") + 1])
                title = str(data.get("title", "")).strip()[:198]
                bullets = [str(b).strip() for b in data.get("bullets", []) if str(b).strip()][:6]
                if bullets and not bullets[-1].lower().startswith("what you get"):
                    bullets.append(f"What You Get: {name} — {warranty}.")
                highlights = [str(h).strip()[:125] for h in data.get("highlights", [])
                              if str(h).strip()][:6]
                if not highlights:
                    highlights = _fallback_highlights(draft)
                return {"title": title, "bullets": bullets, "highlights": highlights,
                        "description": str(data.get("description", "")),
                        "keywords": str(data.get("keywords", ""))}
        except Exception:
            pass

    # ---- Structured rule-based fallback (same shape, keyword-packed) ----
    title = name if (brand and brand.lower() in name.lower()) else f"{brand} {name}".strip()
    # Build keyword phrases: prefer descriptive features (skip raw spec labels like
    # "Rated Voltage"), and synthesise a battery keyword from the parsed mAh.
    bat = _parse_battery(draft.get("feature", ""))
    phrases = []
    if bat.get("mah"):
        phrases.append(f"{int(bat['mah'])}mAh Rechargeable Battery")
    spec_label = re.compile(r"rated|voltage|current|charging input|operating|time|capacity|input",
                            re.I)
    for f in feats:
        if ":" in f:
            lead, val = f.split(":", 1)
            val = val.strip()
            # keep a descriptive value (not a pure number/unit), skip spec labels
            if val and not re.match(r"^[\d.\s/vawhmVAWHM%-]+$", val) and len(val) <= 32:
                phrases.append(val)
        elif 2 < len(f) <= 40 and not spec_label.search(f):
            phrases.append(f)
    for e in dict.fromkeys(phrases):
        cand = f"{title}, {e}"
        if len(cand) < 198:
            title = cand
        else:
            break
    if color and color.lower() not in title.lower() and len(title) + len(color) + 4 < 198:
        title = f"{title}, {color}"
    title = title[:198]

    bullets = []
    for f in feats[:5]:
        if ":" in f:
            lead, desc = f.split(":", 1)
            bullets.append(f"{lead.strip()}: {desc.strip()}")
        else:
            bullets.append(f.strip())
    bullets = [b for b in bullets if b] or [f"{name} by {brand}".strip()]
    bullets.append(f"What You Get: {name} — {warranty}.")

    desc_bits = feats[:12] or [name]
    description = (f"{name}. " + " ".join(desc_bits))[:1900]
    words = re.findall(r"[A-Za-z]{4,}", " ".join([name] + feats[:10]))
    keywords = ", ".join(list(dict.fromkeys(w.lower() for w in words))[:18]) if words else ""
    return {"title": title, "bullets": bullets, "highlights": _fallback_highlights(draft),
            "description": description, "keywords": keywords}


def _common_prefix(skus: list) -> str:
    if not skus:
        return "PARENT"
    p = str(skus[0])
    for s in skus[1:]:
        s = str(s)
        while p and not s.startswith(p):
            p = p[:-1]
    return (p.rstrip("-_ ") or str(skus[0]))[:30]


def _build_feed_items(ready: list) -> tuple:
    """Turn ready 'prepared' items into feed items, grouping colour variants into
    parent+child variation families. Returns (items, family_count)."""
    from collections import defaultdict
    groups = defaultdict(list)
    for p in ready:
        groups[p["draft"].get("group_id", p["sku"])].append(p)
    items, families = [], 0
    for members in groups.values():
        if len(members) == 1:
            p = members[0]
            items.append({"sku": p["sku"], "product_type": p["pt"], "attributes": p["attrs"]})
            continue
        families += 1
        base = members[0]
        prefix = _common_prefix([m["sku"] for m in members]).rstrip("-_") or base["sku"]
        parent_sku = prefix + "-PARENT"
        # Use COLOR if the type defines `color` (COLOR_NAME is deprecated), else COLOR_NAME.
        color_attr = "color" if "color" in base["attrs"] else "color_name"
        theme = "COLOR" if color_attr == "color" else "COLOR_NAME"
        # Parent carries ALL required catalogue attributes (no offer, no single colour).
        parent_flat = dict(base["attrs"])
        for k in ("color", "color_name", "external_product_id", "external_product_id_type",
                  "standard_price", "parent_sku"):
            parent_flat.pop(k, None)
        parent_flat["parentage_level"] = "parent"
        parent_flat["variation_theme"] = theme
        items.append({"sku": parent_sku, "product_type": base["pt"], "attributes": parent_flat})
        for m in members:
            child = dict(m["attrs"])
            child["parentage_level"] = "child"
            child["variation_theme"] = theme
            child["parent_sku"] = parent_sku           # → child_parent_sku_relationship
            child[color_attr] = m["draft"].get("color", "") or child.get(color_attr, "")
            items.append({"sku": m["sku"], "product_type": m["pt"], "attributes": child})
    return items, families


_BADGE_MAX_BYTES = 50000   # badges are tiny (~2-7 KB); product photos are 100 KB+


def _detect_row_badges(xlsx_bytes: bytes) -> tuple:
    """Read embedded BADGE images from an .xlsx and map each to its row. A row can
    hold several images (product photos + a small badge); the badge is the SMALL,
    repeated image — large product photos are ignored.
    Returns ({row_index_0based: image_md5}, {image_md5: image_bytes})."""
    import zipfile
    import hashlib
    import posixpath
    from collections import defaultdict, Counter
    from xml.etree import ElementTree as ET
    SD = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    A = "http://schemas.openxmlformats.org/drawingml/2006/main"
    R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    row_hash, hash_bytes = {}, {}
    try:
        z = zipfile.ZipFile(io.BytesIO(xlsx_bytes))
        names = z.namelist()
        media = {}  # path -> (md5, bytes)
        for n in names:
            if n.startswith("xl/media/"):
                b = z.read(n)
                media[n] = (hashlib.md5(b).hexdigest(), b)
        row_imgs = defaultdict(list)  # row -> [(hash, size, bytes)]
        for dn in [n for n in names if re.match(r"xl/drawings/drawing\d+\.xml$", n)]:
            rels = "xl/drawings/_rels/" + dn.split("/")[-1] + ".rels"
            rid2m = {}
            if rels in names:
                for rel in ET.fromstring(z.read(rels)):
                    t = rel.get("Target", "")
                    rid2m[rel.get("Id")] = (t.lstrip("/") if t.startswith("/")
                                            else posixpath.normpath(posixpath.join("xl/drawings", t)))
            for anchor in ET.fromstring(z.read(dn)):
                frm = anchor.find(f"{{{SD}}}from")
                blip = anchor.find(f".//{{{A}}}blip")
                if frm is None or blip is None:
                    continue
                row = int(frm.find(f"{{{SD}}}row").text)
                m = rid2m.get(blip.get(f"{{{R}}}embed"))
                if m in media:
                    h, b = media[m]
                    row_imgs[row].append((h, len(b), b))
        # Badge = a SMALL image; when a row has several, prefer the one whose image
        # repeats across the most rows (the shared badge), then the smallest.
        freq = Counter(h for v in row_imgs.values()
                       for (h, sz, _) in v if sz <= _BADGE_MAX_BYTES)
        for row, imgs in row_imgs.items():
            cands = [(h, sz, b) for (h, sz, b) in imgs if sz <= _BADGE_MAX_BYTES]
            if not cands:
                continue
            cands.sort(key=lambda x: (-freq[x[0]], x[1]))
            h, sz, b = cands[0]
            row_hash[row] = h
            hash_bytes[h] = b
    except Exception:
        pass
    return row_hash, hash_bytes


def _badge_media_type(b: bytes) -> str:
    if b[:8] == b"\x89PNG\r\n\x1a\n":
        return "image/png"
    if b[:3] == b"\xff\xd8\xff":
        return "image/jpeg"
    if b[:6] in (b"GIF87a", b"GIF89a"):
        return "image/gif"
    if b[:4] == b"RIFF" and b[8:12] == b"WEBP":
        return "image/webp"
    return ""


def _classify_badge(img_bytes: bytes, badge_hash: str):
    """Claude-vision classify a badge image → True (New Arrival) / False (other) /
    None (unknown). Cached per image fingerprint."""
    cache = st.session_state.setdefault("aic_badge_class", {})
    if badge_hash in cache:
        return cache[badge_hash]
    res = None
    mt = _badge_media_type(img_bytes or b"")
    key = db.get_setting("anthropic_api_key", "")
    if key and mt:
        try:
            import anthropic
            import base64
            from core import assistant
            c = anthropic.Anthropic(api_key=key)
            b64 = base64.b64encode(img_bytes).decode()
            r = c.messages.create(
                model=assistant.model(), max_tokens=80,
                messages=[{"role": "user", "content": [
                    {"type": "image", "source": {"type": "base64", "media_type": mt, "data": b64}},
                    {"type": "text", "text": "This is a small product badge/sticker from a price "
                     "list. Return STRICT JSON {\"is_new_arrival\": true|false}. true ONLY if it "
                     "indicates a NEW ARRIVAL / new product. A price/sale/discount/'new price' "
                     "badge is false."}]}])
            txt = "".join(x.text for x in r.content if hasattr(x, "text"))
            m = re.search(r'"is_new_arrival"\s*:\s*(true|false)', txt, re.I)
            if m:
                res = (m.group(1).lower() == "true")
        except Exception:
            res = None
    cache[badge_hash] = res
    return res


def _uploads_to_texts(files) -> list:
    """Extract text from a list of uploaded packaging/manual files (PDF/TXT/CSV/XLSX)."""
    out = []
    for f in (files or []):
        nl = f.name.lower()
        b = f.getvalue()
        try:
            if nl.endswith(".pdf"):
                t = _pdf_text(b)
            elif nl.endswith(".txt"):
                t = b.decode("utf-8", "replace")
            elif nl.endswith(".csv"):
                t = pd.read_csv(io.BytesIO(b), dtype=str).to_csv(index=False)
            elif nl.endswith(".xlsx"):
                t = "\n".join(d.to_csv(index=False)
                              for d in pd.read_excel(io.BytesIO(b), sheet_name=None, dtype=str).values())
            else:
                t = ""
        except Exception:
            t = ""
        if t.strip():
            out.append(t)
    return out


def _build_draft_from_row(r: dict, images: list, img_ok: bool,
                          pack_texts: list, man_texts: list) -> dict:
    """Build one draft from a parsed price-list row + this item's packaging/manual
    text (already extracted). Runs all the spec parsers on the combined text."""
    sku = str(r.get("sku") or "").strip()
    feature = str(r.get("feature") or "")
    bc_raw = str(r.get("barcode") or "").strip()
    pack_text = "\n".join(pack_texts)
    pack_dims = _parse_dimensions(pack_text) if pack_text else {}
    if pack_text:
        feature = feature + "\n" + pack_text[:1500]
    man_text = "\n".join(man_texts)
    if man_text:
        feature = feature + "\n" + man_text[:2500]
        if not pack_dims:
            pack_dims = _parse_dimensions(man_text)
    extra = {}
    if str(r.get("color") or "").strip():
        extra["color"] = r["color"]
    if bc_raw:
        bc = _normalize_barcode(bc_raw)
        extra["external_product_id"] = bc
        extra["external_product_id_type"] = "UPC" if len(bc) == 12 else "EAN"
    bat = _parse_battery(feature)
    if bat.get("mah"):
        extra["lithium_energy"] = bat["mah"]
        extra["energy_is_mah"] = True
    elif bat.get("lithium_energy"):
        extra["lithium_energy"] = bat["lithium_energy"]
        extra["energy_is_mah"] = False
    mat = _parse_material(feature)
    if mat:
        extra["material"] = mat
    extra.update(_parse_power(feature))
    extra.update(_parse_extra_fields(feature))
    extra.update(pack_dims)
    for k, v in _parse_carton(str(r.get("carton") or "")).items():
        extra.setdefault(k, v)
    return {"sku": sku, "title": str(r.get("title") or ""), "price": r.get("price"),
            "images": images, "img_ok": img_ok, "brand": "", "feature": feature,
            "extra": extra, "group_id": r.get("group_id", 0),
            "is_child": bool(r.get("is_child", False)), "color": str(r.get("color") or "")}


def _fetch_drafts() -> list[dict]:
    """Mode A (price list — primary) / Mode B (URL) → list of draft dicts."""
    mode = st.radio("Input mode", ["📄 Price List (recommended for new arrivals)",
                                   "🔗 Paste URL(s)"], horizontal=True)
    drafts: list[dict] = []
    if mode.startswith("📄"):
        st.caption("Upload your price list (any layout — headers can be lower in the sheet, "
                   "section rows are skipped). RRP is used as the price; the Feature column "
                   "becomes the title/bullets/description; barcodes apply your 0/last-digit rule; "
                   "media hyperlinks are read automatically.")
        up = st.file_uploader("Price list (CSV/Excel)", type=["csv", "xlsx"], key="aic_pricelist")
        use_sample = st.checkbox("Use sample price list", value=not bool(up))
        if not (up or use_sample):
            return []

        # ── Phase 1: parse the list + fetch images ONCE (cached by file identity) ──
        pl_sig = (up.name, up.size) if up else ("sample", use_sample)
        base = st.session_state.get("aic_base")
        if not base or base.get("sig") != pl_sig:
            with st.spinner("Reading price list and fetching product images…"):
                df = _load_pricelist(up, use_sample)
                records, imgs_by_sku, imgok_by_sku = [], {}, {}
                # Detect embedded badge images and map them to rows (xlsx only),
                # then classify each DISTINCT badge once (New Arrival vs other).
                row_badge, badge_bytes = ({}, {})
                if up and up.name.lower().endswith(".xlsx"):
                    row_badge, badge_bytes = _detect_row_badges(up.getvalue())
                    for h in set(row_badge.values()):
                        _classify_badge(badge_bytes.get(h, b""), h)
                for _, r in df.iterrows():
                    sku = str(r.get("sku") or "").strip()
                    media = str(r.get("media") or "")
                    res = (oskar_source.fetch_images_from_media_link(media, sku)
                           if (media or sku) else {"images": [], "ok": False})
                    if not res["images"] and str(r.get("image") or "").startswith("http"):
                        res = {"images": [r["image"]], "ok": True}
                    imgs_by_sku[sku] = res["images"]
                    imgok_by_sku[sku] = res["ok"]
                    rec = {k: r.get(k) for k in r.index}
                    # Find a badge on this row (allow ±1 row drift in the anchor).
                    rr = rec.get("_row")
                    bh = None
                    if rr is not None:
                        for off in (0, 1, -1):
                            if (rr + off) in row_badge:
                                bh = row_badge[rr + off]
                                break
                    rec["badge_hash"] = bh
                    records.append(rec)
                base = {"sig": pl_sig, "records": records, "imgs": imgs_by_sku,
                        "imgok": imgok_by_sku, "badge_bytes": badge_bytes,
                        "has_badges": bool(row_badge)}
                st.session_state["aic_base"] = base
        records = base["records"]
        st.caption(f"Parsed {len(records)} product(s); images fetched from connect.oskarme.com.")

        # ── Phase 2: per-item packaging + manual (attach the exact file to each SKU) ──
        st.markdown(section_label("📦 Packaging & manual per item (optional)"),
                    unsafe_allow_html=True)
        st.caption("Attach each item's own packaging/manual PDF — used for that item's "
                   "dimensions, weights, material and deep specs. Tick 🗑️ on the right to "
                   "exclude an item from this batch (e.g. a duplicate you don't want to create).")
        with st.expander("Global fallback files (matched to items by SKU/barcode in the file)"):
            pack = st.file_uploader("Packaging files (PDF/CSV/Excel)", type=["pdf", "csv", "xlsx"],
                                    accept_multiple_files=True, key="aic_packaging")
            manual = st.file_uploader("User manuals (PDF/TXT/CSV)", type=["pdf", "txt", "csv"],
                                      accept_multiple_files=True, key="aic_manual")
        global_pack = _load_packaging(pack)
        global_man = _load_packaging(manual)

        # Keyed by ROW INDEX (not SKU): a file may contain duplicate SKUs, and two
        # widgets with the same key crash Streamlit (StreamlitDuplicateElementKey).
        # `removed` = row indices the user ticked 🗑️ to exclude from this batch.
        per_pkg, per_man, removed = {}, {}, set()
        for idx, rec in enumerate(records):
            sku = str(rec.get("sku") or "").strip()
            if not sku:
                continue
            title = str(rec.get("title") or "")[:46]
            c = st.columns([3, 2, 2, 1])
            # 🗑️ exclude this item from creation (reversible — untick to bring it back).
            if c[3].checkbox("🗑️", key=f"aic_rm_{idx}_{sku}",
                             help="Remove this item from this creation batch"):
                removed.add(idx)
                c[0].markdown(f"~~{title}~~")
                c[0].caption(f"⛔ {sku} excluded")
                continue   # no uploaders for an excluded item
            c[0].markdown(f"**{title}**")
            c[0].code(sku, language=None)   # native one-click copy button

            per_pkg[idx] = c[1].file_uploader(
                "📦 Packaging", type=["pdf", "txt", "csv", "xlsx"], accept_multiple_files=True,
                key=f"aic_pkg_{idx}_{sku}", label_visibility="collapsed")
            per_man[idx] = c[2].file_uploader(
                "📄 Manual", type=["pdf", "txt", "csv"], accept_multiple_files=True,
                key=f"aic_man_{idx}_{sku}", label_visibility="collapsed")
        if removed:
            st.caption(f"⛔ {len(removed)} item(s) excluded — they won't be created. "
                       "Untick 🗑️ to restore.")

        # ── Phase 3: build drafts (per-item files first, else the global fallback) ──
        for idx, rec in enumerate(records):
            if idx in removed:
                continue   # user ticked 🗑️ to exclude this item from the batch
            sku = str(rec.get("sku") or "").strip()
            bc_raw = str(rec.get("barcode") or "").strip()
            pack_texts = _uploads_to_texts(per_pkg.get(idx))
            man_texts = _uploads_to_texts(per_man.get(idx))
            if not pack_texts:
                g = _packaging_for(sku, bc_raw, global_pack)
                if g:
                    pack_texts = [g]
            if not man_texts:
                g = _packaging_for(sku, bc_raw, global_man)
                if g:
                    man_texts = [g]
            dr = _build_draft_from_row(
                rec, base["imgs"].get(sku, []), base["imgok"].get(sku, False),
                pack_texts, man_texts)
            dr["badge_hash"] = rec.get("badge_hash")
            drafts.append(dr)
        st.session_state["aic_drafts"] = drafts
    else:
        st.caption("Paste one item URL per line — the scraper reads Open Graph / JSON-LD / "
                   "common selectors for title, price, brand, description and images.")
        urls = st.text_area("Item URL(s)", placeholder="https://www.example.com/product/123",
                            height=90)
        if st.button("🔎 Fetch from URL(s)") and urls.strip():
            got = []
            for url in [u.strip() for u in urls.splitlines() if u.strip()]:
                res = oskar_source.scrape_product_from_url(url)
                got.append({"sku": "", "title": res.get("title", ""),
                            "price": res.get("price"), "images": res.get("images", []),
                            "img_ok": bool(res.get("images")), "brand": res.get("brand", ""),
                            "source_url": url, "extra": {}})
            st.session_state["aic_drafts"] = got
    return st.session_state.get("aic_drafts", drafts)


def _auto_attributes(draft: dict, reqs: list, opt: dict) -> dict:
    """Build the full attribute dict from price-list data + defaults (no user input)."""
    a = {f["name"]: _prefill(f, draft, opt) for f in reqs}
    return _enrich_attributes(a, draft, opt)


# Item (product) dimensions are NEVER entered by hand — they're auto-derived from
# the package in _enrich_attributes (item fits inside its box). Drop every ITEM
# dimension field from the rendered form: the per-axis ones AND the combined
# "Item Dimensions L x W x H" attribute (item_dimensions / item_length_width_height).
# PACKAGE dimension fields are kept — the user only ever fills the box size.
_ITEM_DIM_FIELDS = {"item_length", "item_width", "item_height", "dimension_unit",
                    "item_depth_width_height", "item_dimensions",
                    "item_length_width_height"}


def _strip_item_dim_fields(fields: list) -> list:
    def _is_item_dim(name: str) -> bool:
        n = (name or "").lower()
        if n in _ITEM_DIM_FIELDS:
            return True
        # Combined item-dimension attributes some types name differently. Never
        # strip PACKAGE dimensions or display (screen) dimensions.
        return (n.startswith("item") and ("dimension" in n or "length_width_height" in n)
                and "package" not in n and "display" not in n)
    return [f for f in fields if not _is_item_dim(f.get("name", ""))]


def _req_missing(reqs: list, attrs: dict) -> list:
    """Labels of required fields that are still blank. A numeric 0 counts as blank —
    a 0 dimension/weight/price is not a real value and Amazon rejects it, so this
    keeps 'ready' honest (an item is never shown ready while a required number is 0)."""
    out = []
    for f in reqs:
        if not f.get("required"):
            continue
        s = str(attrs.get(f["name"], "")).strip()
        blank = not s
        if not blank and f.get("type") == "number":
            try:
                blank = float(s) == 0
            except ValueError:
                blank = False
        if blank:
            out.append(f["label"])
    return out


def _created_local(iso_utc: str):
    """A stored UTC ISO timestamp → a LOCAL datetime (None if unparseable), so the
    created-items log shows the date in the seller's own timezone, not UTC."""
    from datetime import datetime
    if not iso_utc:
        return None
    try:
        return datetime.fromisoformat(str(iso_utc)).astimezone()
    except Exception:
        return None


def _cached_copy(draft: dict) -> dict:
    """Cache the (Claude) listing copy per item so it isn't regenerated on every
    Streamlit rerun — only when the title/feature/brand actually change. Big speedup."""
    sig = hash((str(draft.get("title", "")), str(draft.get("feature", "")),
               str(draft.get("brand", ""))))
    key = str(draft.get("sku") or draft.get("title") or "")
    cache = st.session_state.setdefault("aic_copy_cache", {})
    ent = cache.get(key)
    if ent and ent[0] == sig:
        return ent[1]
    cp = _listing_copy(draft)
    cache[key] = (sig, cp)
    return cp


_POLICY_SYSTEM = (
    "You are an Amazon.ae marketplace compliance reviewer for a seller who is the authorized "
    "brand owner / distributor of these HOUSE brands: Green Lion, Porodo, Porodo Lifestyle, "
    "Powerology, LePresso. The seller's own brand on a product is ALWAYS fine — never flag a "
    "house brand, and NEVER flag merely because a brand name appears in the title or because "
    "the title brand differs from a configured default. Given ONE product (title + features), "
    "decide if it is safe to auto-list or should be FLAGGED for a REAL Amazon policy problem. "
    "Ordinary consumer accessories (cases, holders, cables, chargers, fans, organizers, "
    "speakers, car accessories) are normally OK. Only flag genuine risks. Flag categories: "
    "'Restricted product' (weapons, knives marketed for self-defense, certain medical/health "
    "devices, hazardous goods), 'Trademark / IP' (uses a THIRD-PARTY brand's logo/character/"
    "design without authorization — e.g. Apple, Samsung, Nike, Disney — or a counterfeit "
    "look-alike; do NOT flag the seller's own house brands), 'Category approval' (needs Amazon "
    "approval/gating), 'Prohibited content', 'Safety / compliance'. Return STRICT JSON only: "
    '{"status":"ok"|"flag","flag":"<one category or empty>","reason":"<short why, or empty>"}')


def _policy_check(draft: dict) -> dict:
    """AI compliance gate per item → {'status','flag','reason'}. Cached per title.
    Falls back to 'ok' (with a note) if no Claude key, so it never blocks silently
    on a technical failure — only on a real, stated policy concern."""
    key = str(draft.get("sku") or draft.get("title") or "")
    cache = st.session_state.setdefault("aic_policy_cache", {})
    sig = hash((str(draft.get("title", "")), str(draft.get("feature", ""))[:400], 2))
    ent = cache.get(key)
    if ent and ent[0] == sig:
        return ent[1]
    if not db.get_setting("anthropic_api_key", ""):
        res = {"status": "ok", "flag": "", "reason": "(no AI key — policy check skipped)"}
        cache[key] = (sig, res)
        return res
    try:
        from core import assistant
        prompt = (f"Product: {draft.get('title','')}\nBrand: {draft.get('brand','')}\n"
                  f"Features:\n{draft.get('feature','')[:1200]}")
        txt, status = assistant.complete(_POLICY_SYSTEM, prompt, max_tokens=300)
        data = None
        if status == "ok" and txt and "{" in txt:
            data = json.loads(txt[txt.find("{"):txt.rfind("}") + 1])
        res = ({"status": data.get("status", "ok"), "flag": data.get("flag", ""),
                "reason": data.get("reason", "")} if data
               else {"status": "ok", "flag": "", "reason": ""})
    except Exception:
        res = {"status": "ok", "flag": "", "reason": "(policy check error)"}
    cache[key] = (sig, res)
    return res


def _catalog_status(sku: str, barcode: str, live: bool) -> dict:
    """Is this EXACT SKU already listed in YOUR Amazon inventory? Only your own
    listing (getListingsItem by your SKU) counts as 'exists' → restock. A barcode
    match to Amazon's PUBLIC catalogue is kept as 'catalog_asin' for reference
    (a different COLOUR you already sell will match here) but does NOT turn a NEW
    colour variant into a restock. Cached per SKU.
    {'exists','asin','by','catalog_asin'}."""
    sku = (sku or "").strip()
    if not live or not sku:
        return {"exists": False, "asin": "", "by": "", "catalog_asin": ""}
    cache = st.session_state.setdefault("aic_catalog_cache", {})
    if sku in cache:
        return cache[sku]
    out = {"exists": False, "asin": "", "by": "", "catalog_asin": ""}
    try:
        info = client().confirm_listing(sku)
        if info.get("exists"):
            out = {"exists": True, "asin": info.get("asin", ""), "by": "sku",
                   "catalog_asin": info.get("asin", "")}
    except Exception:
        pass
    # Barcode → public-catalogue ASIN: REFERENCE ONLY (helps later 'sell on existing
    # ASIN'); never marks the item as already in your inventory.
    if barcode:
        try:
            items = client().search_catalog_items(str(barcode).strip(), "EAN")
            if items:
                out["catalog_asin"] = items[0].get("asin", "")
        except Exception:
            pass
    cache[sku] = out
    return out


def _apply_image_override(draft: dict) -> dict:
    """If the user pasted manual image URLs for this SKU (e.g. when oskar has no
    images), use those instead of the auto-fetched ones. Auto-fetch stays the
    default for every other item."""
    ov = st.session_state.get("aic_img_override", {})
    urls = ov.get(str(draft.get("sku", "")).strip())
    if urls:
        d = dict(draft)
        d["images"] = list(urls)
        return d
    return draft


def _enrich_attributes(a: dict, draft: dict, opt: dict) -> dict:
    """Add cross-cutting attributes that aren't tied to one schema field: ALL
    product images, backend keywords, package quantity and FBA fulfilment."""
    imgs = [u for u in (draft.get("images") or []) if str(u).startswith("http")]
    if imgs:
        a.setdefault("main_image_url", imgs[0])
        a["other_image_urls"] = imgs[1:9]          # up to 8 additional images
    b1, b2 = _split_keywords(_augment_keywords(draft, opt.get("keywords", "")))
    if b1 and not str(a.get("generic_keyword", "")).strip():
        a["generic_keyword"] = b1
    if b2:                       # secondary "Add More" keyword block
        a["generic_keyword_more"] = b2
    if not str(a.get("item_package_quantity", "")).strip():
        a["item_package_quantity"] = 1
    if not str(a.get("fulfillment_channel", "")).strip():
        a["fulfillment_channel"] = "FBA"
    # Power units (wattage_unit/voltage_unit) aren't standalone form fields — carry
    # them from the parsed specs so wattage/voltage map as {value, unit}.
    ex = draft.get("extra", {})
    for u in ("wattage_unit", "voltage_unit"):
        if ex.get(u) and not a.get(u):
            a[u] = ex[u]
    # Item (product) dimensions are ALWAYS derived from the PACKAGE (box) size —
    # never entered by hand. The product fits inside its box, so each axis is a
    # fraction (<1) of the package. Re-derived here (not just at parse time) so it
    # also tracks any package-size edit the user makes in the editor.
    pl = a.get("package_length") or ex.get("package_length")
    pw = a.get("package_width") or ex.get("package_width")
    ph = a.get("package_height") or ex.get("package_height")
    pu = a.get("package_dimension_unit") or ex.get("package_dimension_unit") or "centimeters"
    if pl and pw and ph:
        f = _ITEM_FROM_PACKAGE_FACTOR
        a["item_length"] = round(float(pl) * f, 1) or float(pl)
        a["item_width"] = round(float(pw) * f, 1) or float(pw)
        a["item_height"] = round(float(ph) * f, 1) or float(ph)
        a["dimension_unit"] = pu
    return a


def _confirm_with_retry(sku: str, tries: int = 3, delay: float = 4.0) -> dict:
    """Ask Amazon directly whether the listing exists yet (getListingsItem).
    Retries briefly because creation can lag a few seconds behind submission."""
    info = {"exists": False, "status": "", "asin": "", "issues": []}
    for i in range(tries):
        try:
            info = client().confirm_listing(sku)
        except Exception as e:
            info = {"exists": False, "status": "", "asin": "", "issues": [str(e)]}
        if info.get("exists") and (info.get("asin") or info.get("status")):
            return info
        if i < tries - 1:
            time.sleep(delay)
    return info


def _push_one(sku: str, pt: str, attributes: dict, images: list, use_mock: bool) -> dict:
    """Submit one listing, confirm it directly, and persist. Returns the API
    response enriched with the confirmed ASIN/status so the UI can show a clear,
    final result instead of leaving the user in 'still processing' limbo."""
    res = client().push_listing({"sku": sku, "product_type": pt, "attributes": attributes})
    # If Amazon didn't outright reject, confirm the listing item directly — this
    # is authoritative even when the feed report still says IN_PROGRESS.
    if res.get("status") in ("mock_ok", "ok", "submitted"):
        conf = _confirm_with_retry(sku)
        res["confirmed"] = conf
        if conf.get("exists") and conf.get("asin"):
            res["asin"] = conf["asin"]
            res["listing_status"] = conf.get("status", "")
            if res.get("status") == "submitted":
                res["status"] = "ok"   # confirmed created → upgrade from 'submitted'
        # 'listed' ONLY when Amazon confirmed an ASIN; otherwise it's still pending
        # (submitted but unconfirmed) — don't mislabel it as created.
        _created = bool(res.get("asin"))
        db.upsert_catalog_item(sku=sku, asin=res.get("asin", ""),
                               title=attributes.get("item_name", ""),
                               brand=attributes.get("brand", ""), category=pt,
                               status="listed" if _created else "pending")
        db.add_ready_to_list(sku, attributes.get("item_name", ""),
                             float(attributes.get("standard_price") or 0), images,
                             {"product_type": pt, "attributes": attributes})
        db.add_task(f"Verify Amazon listing live: {attributes.get('item_name','')}",
                    f"Pushed via SP-API (SKU {sku}).",
                    module="Inventory & Listing Intake", priority="medium", related_id=sku)
    return res


def _auto_creation_tab() -> None:
    st.markdown(section_label("🤖 Auto Item Creation → Amazon"), unsafe_allow_html=True)
    st.markdown(
        "<p style='color:var(--muted); font-size:.85rem; margin-top:-4px'>"
        "Upload your price list of new arrivals → the dashboard auto-fills <b>Amazon's required "
        "fields</b> per item and pushes the complete ones to Amazon via SP-API in one batch.</p>",
        unsafe_allow_html=True)

    use_mock = db.get_setting("use_mock_amazon", "1") == "1"
    bd = _brand_defaults()
    st.markdown(
        badge("Mock Amazon — add SP-API keys in Settings to push for real" if use_mock
              else "Live SP-API", "amber" if use_mock else "green") + " " +
        (badge(f"Brand owner: {bd['brand'] or '(set brand in Settings)'} · GTIN-exempt", "blue")
         if bd["brand_owner"] else ""),
        unsafe_allow_html=True)

    drafts = _fetch_drafts()
    if not drafts:
        return
    drafts = [_apply_defaults(d) for d in drafts]

    # Per-render requirements cache (avoids duplicate live API calls per product type).
    req_cache: dict = {}

    live_mode = db.get_setting("use_mock_amazon", "1") != "1"

    def _has_battery(d: dict) -> bool:
        return _draft_has_battery(d)

    def _valid_product_types() -> set:
        """Cached set of product types that actually exist in this marketplace,
        so a wrong guess is caught before it fails at the feed."""
        if "aic_valid_pts" not in st.session_state:
            try:
                st.session_state["aic_valid_pts"] = set(client().get_product_types())
            except Exception:
                st.session_state["aic_valid_pts"] = set()
        return st.session_state["aic_valid_pts"]

    def _reqs(pt: str, has_battery: bool = False) -> list:
        ck = (pt, has_battery)
        if ck not in req_cache:
            try:
                if live_mode:
                    req_cache[ck] = _relax_for_brand_owner(_hybrid_requirements(pt, has_battery))
                else:
                    req_cache[ck] = _relax_for_brand_owner(client().get_listing_requirements(pt))
            except Exception:
                req_cache[ck] = _relax_for_brand_owner(_core_fields()) if live_mode else []
        # Item dimensions are auto-derived from the package, so they're never shown
        # as form fields (the user only fills the package size).
        return _strip_item_dim_fields(req_cache[ck])

    # ---- BATCH overview --------------------------------------------------
    st.markdown(section_label("1 · Batch — validate all items"), unsafe_allow_html=True)
    prepared = []
    rows = []
    valid_pts = _valid_product_types()
    pt_overrides = st.session_state.setdefault("aic_pt_override", {})

    # Policy + catalog scan (explicit, cached → no per-render API/AI calls).
    sc = st.columns([1, 2])
    if sc[0].button("🛡️ Run policy + catalog scan", use_container_width=True):
        with st.spinner("Checking Amazon policies and catalogue existence…"):
            for d in drafts:
                s = (d.get("sku") or _slug(d.get("title", ""))).strip()
                _policy_check(d)
                cat = _catalog_status(s, str(d.get("extra", {}).get("external_product_id") or ""),
                                      live_mode)
                if cat["exists"]:
                    db.add_restock(s, d.get("title", ""), cat.get("asin", ""))
        st.session_state["aic_scanned"] = True
        st.rerun()
    sc[1].caption("Flags Amazon policy risks (restricted / trademark / approval) and detects items "
                  "already in your catalogue → routed to the Restock tab. Runs once; cached.")
    scanned = st.session_state.get("aic_scanned", False)
    policy_ovr = st.session_state.setdefault("aic_policy_override", set())
    restock_skus = {r["sku"] for r in db.get_restock() if r.get("status") != "done"}
    routed = []

    # RESTOCK vs NEW ARRIVAL is decided by YOUR Amazon inventory — NOT by badges.
    # Each SKU is looked up in Manage All Inventory (getListingsItem): found → you
    # already sell it → Restock; not found → it's a new arrival → created here.
    st.caption("🔎 Each SKU is looked up in your Amazon inventory: **found → Restock**, "
               "**not found → new arrival** (created here).")

    for i, d in enumerate(drafts):
        d = _apply_image_override(d)        # use manual image URLs if provided
        sku = d.get("sku") or _slug(d.get("title", ""))
        # The ONLY signal: is this exact SKU already in your Amazon inventory?
        # Found → restock; not found → new arrival → create here. A stale entry in
        # the restock queue does NOT override this live inventory truth (so a SKU you
        # don't actually sell can never be stuck in Restock).
        cat = _catalog_status(sku, str(d.get("extra", {}).get("external_product_id") or ""),
                              live_mode)
        if cat.get("exists"):
            db.add_restock(sku, d.get("title", ""), cat.get("asin", ""))
            routed.append((sku, d.get("title", "")))
            continue
        elif sku in restock_skus:
            # Was queued before but is NOT in your inventory → it's a new arrival;
            # drop the stale queue entry so it isn't stuck in the Restock tab either.
            db.remove_restock(sku)
        # SMART product-type detection. Amazon's OWN product-type search is the
        # authoritative signal — it returns the same category shown on the live
        # listing (e.g. 'car cup holder' → CUP_HOLDER, 'car seat organizer' →
        # CADDY) and isn't fooled by accessory words in the title (a cup-holder
        # whose title mentions a bundled 'charging cable' must NOT become CABLE).
        # The keyword rules are only a fast fallback when the search returns nothing.
        if sku in pt_overrides:
            pt = pt_overrides[sku]
        else:
            pt = _amazon_product_type(d.get("title", ""), valid_pts) \
                or _guess_product_type(d.get("title", ""))
            if valid_pts and pt not in valid_pts:
                pt = "GENERIC"
        opt = _cached_copy(d)
        # GENERIC (or any type not in this marketplace's catalogue) is NOT pushable
        # — it gets rejected at the feed with code 4000003. Make the user pick a
        # valid type in the editor instead of silently failing later.
        if pt == "GENERIC" or (valid_pts and pt not in valid_pts):
            prepared.append({"i": i, "draft": d, "pt": pt, "reqs": [], "attrs": {},
                             "sku": sku, "missing": ["set product type"], "opt": opt})
            rows.append({"SKU": sku, "Title": (d.get("title") or "")[:42], "Type": pt,
                         "Price": "—", "Status": "set product type"})
            continue
        reqs = _reqs(pt, _has_battery(d))
        if not reqs:
            # Couldn't resolve Amazon's fields for the guessed type — fix in editor.
            prepared.append({"i": i, "draft": d, "pt": pt, "reqs": [], "attrs": {},
                             "sku": sku, "missing": ["set product type"], "opt": opt})
            rows.append({"SKU": sku, "Title": (d.get("title") or "")[:42], "Type": pt,
                         "Price": "—", "Status": "set product type"})
            continue
        attrs = _auto_attributes(d, reqs, opt)
        missing = _req_missing(reqs, attrs)
        # If this SKU was completed & saved in the editor, use those finished
        # attributes and treat it as ready (so it joins the variation/batch push).
        saved = st.session_state.get("aic_completed", {}).get(sku)
        if saved:
            attrs = {**attrs, **saved["attrs"]}    # saved edits win over auto-fill
            pt = saved.get("pt", pt)
            if saved.get("images"):
                d = dict(d)
                d["images"] = saved["images"]
            # Re-check against CURRENT requirements — a previously-saved item can be
            # missing a newly-enforced field (e.g. automotive fit), so it must NOT be
            # shown 'ready' just because it was saved earlier.
            missing = _req_missing(reqs, attrs)
        # Policy gate: a flagged item is blocked from push until explicitly overridden.
        policy = _policy_check(d) if scanned else {"status": "?", "flag": "", "reason": ""}
        if policy.get("status") == "flag" and sku not in policy_ovr:
            missing = [f"⛔ {policy.get('flag') or 'policy'}"] + missing
        prepared.append({"i": i, "draft": d, "pt": pt, "reqs": reqs, "attrs": attrs,
                         "sku": sku, "missing": missing, "opt": opt, "policy": policy})
        rows.append({"SKU": sku, "Title": (d.get("title") or "")[:42], "Type": pt,
                     "Price": f"AED {float(attrs.get('standard_price') or 0):,.0f}",
                     "Status": ("✓ saved" if saved else
                                ("Ready" if not missing else f"{len(missing)} missing"))})

    if routed:
        st.info(f"🔁 {len(routed)} item(s) already in your catalogue → moved to the **Restock** "
                f"tab: " + ", ".join(s for s, _ in routed[:8])
                + ("…" if len(routed) > 8 else ""))

    # Compact per-row list with an inline Edit button (jumps to the detailed editor).
    hdr = st.columns([3, 2, 1, 2, 1])
    for col, label in zip(hdr, ("Item (SKU)", "Type", "Price", "Status / issues", "")):
        col.markdown(f"**{label}**")
    for p, row in zip(prepared, rows):
        sku = p["sku"]
        i = p["i"]   # unique per prepared item — widget keys must not collide on dup SKUs
        ready = not p["missing"]
        saved = sku in st.session_state.get("aic_completed", {})
        pol = p.get("policy", {})
        rc = st.columns([3, 2, 1, 2, 1])
        rc[0].markdown(f"{row['Title']}  \n`{sku}`")
        rc[1].markdown(f"`{row['Type']}`")
        rc[2].markdown(row["Price"])
        if pol.get("status") == "flag" and sku not in policy_ovr:
            rc[3].markdown(badge(f"⛔ {pol.get('flag') or 'policy'}", "coral"),
                           unsafe_allow_html=True)
            if pol.get("reason"):
                rc[3].caption(pol["reason"][:90])
            if rc[4].button("Override", key=f"aic_polovr_{i}_{sku}", use_container_width=True,
                            help="Confirm this item complies and allow it to be pushed."):
                policy_ovr.add(sku)
                st.rerun()
            continue
        if saved:
            rc[3].markdown(badge("✓ saved", "green"), unsafe_allow_html=True)
        elif ready:
            ok = "✓ Ready" + (" · policy✔ (override)" if sku in policy_ovr else
                              (" · policy✔" if pol.get("status") == "ok" else ""))
            rc[3].markdown(badge(ok, "green"), unsafe_allow_html=True)
        else:
            msg = ", ".join(p["missing"][:4]) + ("…" if len(p["missing"]) > 4 else "")
            rc[3].markdown(badge(f"⚠ {msg}", "amber"), unsafe_allow_html=True)
        if rc[4].button("✏️ Edit", key=f"aic_edit_btn_{i}_{sku}", use_container_width=True):
            st.session_state["aic_pick"] = p["i"]
            st.session_state["aic_focus_editor"] = True
            st.rerun()

    ready_items = [p for p in prepared if not p["missing"]]
    # Detect colour-variation families among the items.
    from collections import Counter
    gcounts = Counter(p["draft"].get("group_id", p["sku"]) for p in prepared)
    fam_count = sum(1 for n in gcounts.values() if n > 1)
    if fam_count:
        st.caption(f"🎨 {fam_count} colour-variation family(ies) detected — these will be "
                   f"created as a parent listing with colour children.")

    # Persistent result of the LAST push — survives the reruns that previously
    # wiped the success/error message off the screen ("nothing happened").
    last = st.session_state.get("aic_last_push")
    if last:
        (st.success if last.get("ok") else st.warning)(
            "**Last push result:**\n" + "\n".join(f"- {ln}" for ln in last.get("lines", [])))
        if st.button("Dismiss result", key="aic_dismiss_push"):
            st.session_state.pop("aic_last_push", None)
            st.rerun()

    # ── Created-items log: which items were created, WHEN, filterable by date ──
    with st.expander("📅 Created items log — see what was created & filter by date",
                     expanded=False):
        created = db.get_created_log()
        if not created:
            st.caption("Nothing created yet. Every item you push to Amazon is recorded here "
                       "with its creation date.")
        else:
            for r in created:                       # attach a local date/time to each row
                ldt = _created_local(r.get("created_at"))
                r["_day"] = ldt.strftime("%Y-%m-%d") if ldt else ""
                r["_when"] = ldt.strftime("%Y-%m-%d %H:%M") if ldt else (r.get("created_at") or "")
            days = sorted({r["_day"] for r in created if r["_day"]}, reverse=True)
            fc = st.columns([1, 2, 2])
            all_days = fc[0].checkbox("All dates", value=False, key="aic_log_all")
            from datetime import date as _date
            default_day = _date.fromisoformat(days[0]) if days else None
            picked = fc[1].date_input("Show items created on", value=default_day,
                                      key="aic_log_date", disabled=all_days)
            sel = picked.isoformat() if hasattr(picked, "isoformat") else str(picked)
            shown = created if all_days else [r for r in created if r["_day"] == sel]
            fc[2].markdown(
                "<div style='padding-top:28px'>" +
                badge(f"{len(shown)} created" + ("" if all_days else f" on {sel}"),
                      "green" if shown else "amber") + "</div>", unsafe_allow_html=True)
            if shown:
                dfc = pd.DataFrame([{
                    "Created": r["_when"], "SKU": r.get("sku") or "",
                    "Title": (r.get("title") or "")[:48], "ASIN": r.get("asin") or "—",
                    "Category": r.get("category") or "",
                    "Price (AED)": f"{float(r.get('price') or 0):,.0f}"} for r in shown])
                styled_table(dfc)
                export_buttons(dfc, "aic_created_log")
            else:
                st.caption(f"No items created on {sel}. Pick another date or tick 'All dates'.")

    c1, c2 = st.columns([1, 2])
    with c1:
        st.markdown(badge(f"{len(ready_items)}/{len(prepared)} ready to push",
                          "green" if ready_items else "amber"), unsafe_allow_html=True)
    with c2:
        if st.button(f"🚀 Push ALL {len(ready_items)} ready items to Amazon",
                     use_container_width=True, type="primary", disabled=not ready_items):
            log = []   # every line ends up in the persistent result panel above
            ok = False
            try:
                # Build ONE JSON_LISTINGS_FEED, grouping colour variants into families.
                items, fam_n = _build_feed_items(ready_items)
                log.append(f"Built {len(items)} feed message(s) from {len(ready_items)} ready "
                           f"item(s) · {fam_n} variation family(ies).")
                if not items:
                    log.append("⛔ Nothing to submit — no feed messages were built. "
                               "Check that each item has a valid product type and price.")
                    raise RuntimeError("empty feed")
                with st.spinner(f"Submitting JSON_LISTINGS_FEED "
                                f"({len(items)} messages, {fam_n} variation families)…"):
                    res = client().push_listings_feed(items)
                per = res.get("per_sku", {})
                # Confirm each item directly (getListingsItem) so 'submitted' becomes a
                # definite created/ASIN result without leaving the dashboard.
                confirmed_asins = {}
                for p in ready_items:
                    info = per.get(p["sku"], {})
                    if info.get("status") in ("accepted", "submitted"):
                        conf = (_confirm_with_retry(p["sku"], tries=2, delay=3.0)
                                if not use_mock else {"asin": "MOCK-ASIN", "status": "MOCK"})
                        if conf.get("asin"):
                            confirmed_asins[p["sku"]] = conf
                        # 'listed' ONLY when Amazon confirmed an ASIN — an unconfirmed
                        # 'submitted' item stays 'pending' so it isn't shown as created.
                        db.upsert_catalog_item(sku=p["sku"], asin=conf.get("asin", ""),
                                               title=p["attrs"].get("item_name", ""),
                                               brand=p["attrs"].get("brand", ""), category=p["pt"],
                                               status="listed" if conf.get("asin") else "pending")
                        db.add_ready_to_list(p["sku"], p["attrs"].get("item_name", ""),
                                             float(p["attrs"].get("standard_price") or 0),
                                             p["draft"].get("images", []),
                                             {"product_type": p["pt"], "attributes": p["attrs"]})
                sub = res.get("submitted", 0)
                rej = res.get("rejected", 0)
                log.append(f"Feed **{res.get('feedId','—')}** · status "
                           f"{res.get('processingStatus','—')} · accepted "
                           f"{res.get('accepted',0)} / rejected {rej}"
                           + (f" / still-processing {sub}" if sub else "")
                           + f" ({'MOCK' if use_mock else 'LIVE'}).")
                if confirmed_asins:
                    from datetime import datetime
                    _stamp = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M")
                    log.append(f"🕒 Created on **{_stamp}** (local) — see the "
                               "'📅 Created items log' above to filter by date.")
                for s, c in confirmed_asins.items():
                    log.append(f"✅ {s} → ASIN {c['asin']} ({c.get('status') or 'live'})")
                unconfirmed = sub - len(confirmed_asins)
                if unconfirmed > 0:
                    log.append(f"⏳ {unconfirmed} item(s) still processing on Amazon — use "
                               f"'🔄 Re-check status' in the editor below in a minute.")
                rejected = {s: i.get("issues", []) for s, i in per.items()
                            if i.get("status") == "rejected"}
                # Self-heal: read missing field names out of each item's reject
                # messages and pre-add them to that product type's form.
                pt_by_sku = {p["sku"]: p["pt"] for p in ready_items}
                added = False
                for s, iss in rejected.items():
                    log.append(f"❌ {s}: {'; '.join(iss) or 'see Amazon feed report'}")
                    attrs_needed = _attrs_from_messages(iss)
                    ptx = pt_by_sku.get(s)
                    if attrs_needed and ptx:
                        st.session_state.setdefault("aic_more", {}).setdefault(ptx, set()).update(
                            attrs_needed)
                        added = True
                if added:
                    req_cache.clear()
                    log.append("➕ Added the missing field(s) to the form and pre-filled them — "
                               "fix any blanks in the editor below, then push again.")
                ok = rej == 0 and bool(confirmed_asins or sub or res.get("accepted", 0))
            except Exception as e:
                log.append(f"⛔ Push failed: {e}")
            st.session_state["aic_last_push"] = {"ok": ok, "lines": log}
            st.rerun()
    if len(ready_items) < len(prepared):
        st.caption("Items with missing fields → complete them in the detailed editor below.")

    # ---- DETAILED editor (complete / fix one item) ----------------------
    st.markdown(section_label("2 · Detailed editor — complete or fix one item"),
                unsafe_allow_html=True)
    _focus = st.session_state.pop("aic_focus_editor", False)
    if not prepared:
        st.info("No items to edit here — every item in this file was routed to the **Restock** "
                "tab (none carried a 'New Arrival' badge / all already in your catalogue).")
        return
    with st.expander("✏️ Open detailed editor",
                     expanded=_focus or bool(len(ready_items) < len(prepared))):
        labels = [f"{p['i']+1}. {p['draft'].get('title') or '(untitled)'} "
                  f"({'ready' if not p['missing'] else str(len(p['missing']))+' missing'})"
                  for p in prepared]
        sel = st.selectbox("Item", range(len(prepared)), format_func=lambda i: labels[i],
                           key="aic_pick")
        if sel is None:
            sel = 0
        p = prepared[sel]
        draft, pt0 = p["draft"], p["pt"]
        draft = _apply_image_override(draft)   # honour any manual image URLs
        # Type override: search Amazon's catalogue + show the resolved type at the top.
        types = sorted({str(x).strip() for x in (client().get_product_types() or [])
                        if str(x).strip()})
        kw_box = st.columns([3, 2])
        kw = kw_box[0].text_input("🔍 Search a different product type (e.g. 'cup holder')",
                                  value="", key=f"aic_kwsearch_{p['sku']}",
                                  placeholder="leave blank to keep the auto-detected type")
        if kw_box[1].button("Find type", key=f"aic_findbtn_{p['sku']}",
                            use_container_width=True) and kw.strip():
            try:
                raw = client().get_product_types(kw.strip())
                cands = [str(x).strip() for x in (raw or []) if str(x).strip()]
                if cands:
                    pt_overrides[p["sku"]] = cands[0]
                    others = ", ".join(cands[1:6]) or "—"
                    st.success(f"Set to **{cands[0]}** (Amazon match for '{kw.strip()}'). "
                               f"Other matches: {others}")
                    p["pt"] = cands[0]; pt0 = cands[0]
                else:
                    st.warning(f"No Amazon product type matched '{kw.strip()}'.")
            except Exception as e:
                st.warning(f"Search failed: {e}")
        pt = st.selectbox("Amazon product type", types,
                          index=types.index(pt0) if pt0 in types else 0, key="aic_pt")
        if pt != pt0:
            pt_overrides[p["sku"]] = pt
        reqs = _reqs(pt, _has_battery(draft))
        if not reqs:
            st.warning("Couldn't load Amazon's required fields for this product type. "
                       "Pick a different product type above (the list comes from your live "
                       "Amazon marketplace).")
        opt = _cached_copy(draft)
        draft_id = _slug(draft.get("sku") or draft.get("title") or str(sel))

        if draft.get("images"):
            st.image(draft["images"][0], width=130, caption="Product image")

        ov_sku = str(draft.get("sku", "")).strip()
        grp = draft.get("group_id")
        fam = [q for q in prepared if q["draft"].get("group_id") == grp]
        saved_map = st.session_state.get("aic_completed", {})
        if len(fam) > 1:
            sibs = ", ".join(f"{q['draft'].get('color','?')}"
                             f"{'✓' if q['sku'] in saved_map else ''}" for q in fam)
            st.info(f"🎨 **Colour variation family** ({len(fam)} colours: {sibs}). Complete & "
                    f"**Apply & Save each colour**, then click **'🚀 Push ALL'** at the top to "
                    f"create them together as ONE variation. "
                    f"Saved: {len([q for q in fam if q['sku'] in saved_map])}/{len(fam)}.")

        # 'Item Highlight' can't be submitted via the API (Amazon: "unsupported") —
        # show the generated text so it can be pasted into Seller Central manually.
        _hl = [h for h in opt.get("highlights", []) if h]
        if _hl:
            st.info("💡 **Item Highlight** (paste manually in Seller Central — not accepted via "
                    "API; shows only when the title is under 75 chars):\n\n" + " · ".join(_hl))

        seller_sku = st.text_input("Seller SKU *",
                                   value=draft.get("sku") or _slug(draft.get("title", "")),
                                   key=f"aic_sku_{draft_id}")

        # All edits live inside a FORM — nothing reloads/applies until you click the
        # Apply button (no per-keystroke reruns), which also Saves it ready to push.
        with st.form(f"aic_form_{draft_id}", clear_on_submit=False):
            st.caption("Edit everything below, then click **Apply edits & Save** once — the page "
                       "won't reload on each change (much faster).")
            cur_imgs = st.session_state.get("aic_img_override", {}).get(
                ov_sku, draft.get("images") or [])
            img_txt = st.text_area(
                "🖼️ Product image URLs (auto-fetched from oskar; paste public URLs to override "
                "— first line = main image)", value="\n".join(cur_imgs), height=80,
                key=f"aic_imgurls_{draft_id}")
            attributes = {}
            empty_req_keys = []   # required fields left blank → outlined red below

            def _render_field(f, container):
                """Render one schema field into a given column and store its value."""
                key = f"aicf_{draft_id}_{f['name']}"
                prefill = _prefill(f, draft, opt)
                label = f["label"] + (" *" if f.get("required") else "")
                with container:
                    if f["type"] == "select":
                        opts = f["options"] or [""]
                        default = prefill if prefill in opts else f.get("default", opts[0])
                        val = st.selectbox(label, opts,
                                           index=opts.index(default) if default in opts else 0,
                                           key=key)
                    elif f["type"] == "number":
                        val = st.number_input(
                            label, value=_num(prefill, _num(f.get("default", 0))), key=key)
                    elif f["type"] == "textarea":
                        val = st.text_area(label, value=str(prefill or ""), height=90, key=key)
                    else:
                        val = st.text_input(label, value=str(prefill or ""), key=key)
                attributes[f["name"]] = val
                if f.get("required"):
                    blank = (val is None or (isinstance(val, str) and not val.strip())
                             or (f["type"] == "number" and float(val or 0) == 0))
                    if blank:
                        empty_req_keys.append(key)

            # Group related measurement fields so they sit together, each VALUE next
            # to its UNIT on the same row (was a flat 2-col grid that scattered them).
            def _isdim(f):
                s = (f["name"] + " " + f["label"]).lower()
                return any(w in s for w in ("length", "width", "height", "depth", "dimension"))

            def _iswt(f):
                return "weight" in (f["name"] + " " + f["label"]).lower()

            def _isunit(f):
                return "unit" in (f["name"] + " " + f["label"]).lower()

            dim_f = [f for f in reqs if _isdim(f)]
            wt_f = [f for f in reqs if _iswt(f) and not _isdim(f)]
            rest_f = [f for f in reqs if f not in dim_f and f not in wt_f]
            # Dimensions: all the length/width/height values first, the unit at the row end.
            dim_f.sort(key=lambda f: (_isunit(f), f["name"]))
            # Weights: keep each weight value immediately followed by its own unit.
            wt_f.sort(key=lambda f: (f["name"].replace("_unit", ""), _isunit(f)))

            if dim_f:
                st.markdown("**📦 Package dimensions** — enter the box size; the item "
                            "size is auto-estimated just under the box (you don't fill it).")
                for i in range(0, len(dim_f), 4):
                    chunk = dim_f[i:i + 4]
                    for c, f in zip(st.columns(len(chunk)), chunk):
                        _render_field(f, c)
            if wt_f:
                st.markdown("**⚖️ Weights** — each weight next to its unit")
                for i in range(0, len(wt_f), 2):
                    chunk = wt_f[i:i + 2]
                    for c, f in zip(st.columns(len(chunk)), chunk):
                        _render_field(f, c)
            if rest_f:
                if dim_f or wt_f:
                    st.markdown("**Other details**")
                rcols = st.columns(2)
                for k, f in enumerate(rest_f):
                    _render_field(f, rcols[k % 2])
            # Outline every still-empty REQUIRED field in red so nothing required
            # is missed before pushing (updates each time you Apply / the page reruns).
            if empty_req_keys:
                sel = ", ".join(
                    f".st-key-{k} input, .st-key-{k} textarea, "
                    f".st-key-{k} div[data-baseweb='select'] > div" for k in empty_req_keys)
                st.markdown(
                    f"<style>{sel}{{border:2px solid #e7503c !important;border-radius:7px "
                    f"!important;box-shadow:0 0 0 2px rgba(231,80,60,.20) !important;}}</style>",
                    unsafe_allow_html=True)
                st.caption(f"🔴 {len(empty_req_keys)} required field(s) still empty — outlined "
                           f"in red above.")
            submitted = st.form_submit_button("✅ Apply edits & Save (ready to push)",
                                              type="primary", use_container_width=True)

        # Build the final attribute set from the applied values.
        new_imgs = [u.strip() for u in img_txt.splitlines() if u.strip().startswith("http")]
        draft = dict(draft)
        draft["images"] = new_imgs
        attributes = _enrich_attributes(attributes, draft, opt)
        missing = _req_missing(reqs, attributes)
        if not reqs:
            missing = ["Select a valid product type"] + missing
        if not seller_sku.strip():
            missing = ["Seller SKU"] + missing
        total = sum(1 for f in reqs if f.get("required")) + 1
        st.markdown(badge(f"{max(0, total - len(missing))}/{total} required complete",
                          "green" if not missing else "amber"), unsafe_allow_html=True)
        if missing:
            st.markdown(alert("Still required: " + ", ".join(missing), kind="amber", icon="✏️"),
                        unsafe_allow_html=True)

        if submitted:
            if new_imgs:
                st.session_state.setdefault("aic_img_override", {})[ov_sku] = new_imgs
            if missing:
                st.warning("Not saved yet — still missing: " + ", ".join(missing))
            else:
                st.session_state.setdefault("aic_completed", {})[seller_sku.strip()] = {
                    "attrs": dict(attributes), "pt": pt, "images": new_imgs,
                    "group_id": grp, "color": draft.get("color", "")}
                st.success(f"✅ Applied & saved {seller_sku.strip()} — ready to push."
                           + (" Now do the next colour, then '🚀 Push ALL' at the top."
                              if len(fam) > 1 else " Use '🚀 Push ALL' at the top, or push below."))
                st.rerun()

        # Catalog check + validation preview (per the SP-API listing flow).
        ce1, ce2 = st.columns(2)
        with ce1:
            if st.button("🔎 Check if ASIN exists", key=f"aic_cat_{draft_id}"):
                bc = str(attributes.get("external_product_id", "")).strip()
                if bc:
                    items = client().search_catalog_items(
                        bc, attributes.get("external_product_id_type", "EAN"))
                    if items:
                        st.info("Existing match(es): " +
                                ", ".join(i.get("asin", "?") for i in items[:5]) +
                                " — consider listing as an offer on the existing ASIN.")
                    else:
                        st.success("No existing ASIN — this will create a new product.")
                else:
                    st.caption("No barcode to search the catalog.")
        with ce2:
            if st.button("✅ Validate (preview, no commit)", key=f"aic_val_{draft_id}"):
                v = client().validate_preview(seller_sku.strip() or "PREVIEW", pt, attributes)
                if v.get("ok"):
                    st.success(f"✅ Valid ({v.get('status')}) — ready to push to Amazon.")
                else:
                    # Auto-discover any required fields Amazon flagged and add them to the form.
                    flagged = set()
                    for i in v.get("issues", []) + v.get("errors", []):
                        if isinstance(i, dict) and i.get("severity", "ERROR") == "ERROR":
                            for an in (i.get("attributeNames") or []):
                                flagged.add(an)
                    flagged |= _attrs_from_messages(v.get("issues", []) + v.get("errors", []))
                    known = {f["name"] for f in reqs}
                    nested_parents = {p for p, subs in _NESTED_FIELDS.items()
                                      if any(s["name"] in known for s in subs)}
                    new = [a for a in flagged if a not in known and a not in nested_parents
                           and a not in _CORE_COVERED]
                    if new:
                        store = st.session_state.setdefault("aic_more", {}).setdefault(pt, set())
                        store.update(new)
                        req_cache.clear()  # rebuild form with the new fields
                        st.warning("Amazon needs more fields: " +
                                   ", ".join(a.replace('_', ' ') for a in new) +
                                   ". I've added them below — fill any blanks and validate again.")
                        st.rerun()
                    else:
                        errs = v.get("errors", []) or v.get("issues", [])
                        msg = "\n".join("• " + (e.get("message", "") if isinstance(e, dict) else str(e))
                                        for e in errs[:10])
                        st.error(("Validation issues:\n" + msg) if msg else f"Invalid: {v.get('status')}")

        if st.button("🚀 Push this item to Amazon (Feeds API)", use_container_width=True,
                     type="primary", disabled=bool(missing)):
            res = _push_one(seller_sku.strip(), pt, attributes, draft.get("images", []), use_mock)
            status = res.get("status")
            fid = res.get("feedId", "—")
            conf = res.get("confirmed", {})
            if status in ("mock_ok", "ok"):
                if conf.get("asin"):
                    st.success(f"✅ CREATED on Amazon — ASIN **{conf['asin']}** · status "
                               f"{conf.get('status') or 'live'} · SKU {seller_sku.strip()}.")
                else:
                    st.success(f"✅ ACCEPTED by Amazon — feed {fid} processed (DONE). "
                               f"SKU {seller_sku.strip()}. It will appear in Seller Central shortly.")
                st.balloons()
            elif status == "submitted":
                st.warning(f"📤 Submitted (feed {fid}) but Amazon hasn't finished processing yet "
                           f"and the listing isn't visible via the API at this second. This usually "
                           f"means it's still in Amazon's queue — click below to re-check.")
                st.session_state["aic_last_push_sku"] = seller_sku.strip()
            elif status == "invalid":
                st.error("Missing required fields: " + ", ".join(res.get("missing", [])))
            else:
                # Auto-discover any fields Amazon's feed report flagged, add to the form.
                report = (res.get("raw") or {}).get("report") or {}
                flagged = set()
                for i in report.get("issues", []):
                    if i.get("severity") == "ERROR":
                        for an in (i.get("attributeNames") or []):
                            flagged.add(an)
                issues = res.get("issues") or []
                # Also read the field title out of the message text (Amazon often
                # omits the structured attribute name) so we self-heal.
                flagged |= _attrs_from_messages(report.get("issues", []) + list(issues))
                known = {f["name"] for f in reqs}
                nested_parents = {p for p, subs in _NESTED_FIELDS.items()
                                  if any(s["name"] in known for s in subs)}
                new = [a for a in flagged if a not in known and a not in nested_parents
                       and a not in _CORE_COVERED]
                if new:
                    st.session_state.setdefault("aic_more", {}).setdefault(pt, set()).update(new)
                    req_cache.clear()
                    st.warning("❌ Amazon rejected — and asked for more fields: " +
                               ", ".join(a.replace('_', ' ') for a in new) +
                               ". Added below — fill any blanks and push again.")
                    st.rerun()
                else:
                    st.error("❌ REJECTED by Amazon:\n" +
                             ("\n".join("• " + str(i) for i in issues) if issues else str(res)))

        # Re-check a SKU that was 'submitted' (feed still processing) — confirms
        # creation directly so the whole flow stays inside the dashboard.
        recheck_sku = st.session_state.get("aic_last_push_sku") or seller_sku.strip()
        if recheck_sku and not use_mock:
            if st.button(f"🔄 Re-check status for {recheck_sku}", key=f"aic_recheck_{draft_id}"):
                conf = _confirm_with_retry(recheck_sku, tries=2, delay=3.0)
                if conf.get("exists") and conf.get("asin"):
                    db.upsert_catalog_item(sku=recheck_sku, asin=conf["asin"],
                                           title=attributes.get("item_name", ""),
                                           brand=attributes.get("brand", ""), category=pt,
                                           status="listed")
                    st.success(f"✅ CREATED — ASIN **{conf['asin']}** · status "
                               f"{conf.get('status') or 'live'}.")
                    st.balloons()
                elif conf.get("issues"):
                    st.error("Amazon reported issues:\n" +
                             "\n".join("• " + str(i) for i in conf["issues"]))
                else:
                    st.info("Still not visible via the API — Amazon is likely still processing. "
                            "Give it another minute and re-check.")

    # ---- submitted / ready queue ----------------------------------------
    st.markdown(section_label("📦 Submitted / Ready-to-List"), unsafe_allow_html=True)
    ready = db.get_ready_to_list()
    if ready:
        def _when(r):
            ldt = _created_local(r.get("created_at"))
            return ldt.strftime("%Y-%m-%d %H:%M") if ldt else (r.get("created_at") or "—")

        def _is_created(r) -> bool:
            # Truth = Amazon gave it a real ASIN. No ASIN → submitted but not created.
            asin = str(r.get("asin") or "").strip()
            return bool(asin) and asin.upper() not in ("PENDING", "")

        uncreated = [r for r in ready if not _is_created(r)]
        n_ok = len(ready) - len(uncreated)
        # Per-SKU issues + live Amazon status from the last health-check.
        issues_map = st.session_state.get("aic_issues", {})
        status_map = st.session_state.get("aic_status_map", {})

        def _issue_cell(sku):
            if sku not in issues_map:
                return "— (not checked)"
            iss = issues_map[sku]
            if not iss:
                return "✓ complete"
            errs = sum(1 for i in iss if i.get("severity") == "ERROR")
            warns = len(iss) - errs
            return f"⚠ {len(iss)} ({errs} block / {warns} info)" if errs else f"⚠ {warns} missing"

        def _status_cell(r):
            # Prefer the real Amazon status from a health-check; else the local
            # created/not-confirmed signal until one has been run.
            s = status_map.get(r["sku"])
            if s:
                return s
            return "Created" if _is_created(r) else "Not confirmed"

        def _status_kind(s: str) -> str:
            t = s.lower()
            if "active" in t or t == "created":
                return "green"
            if any(w in t for w in ("suppress", "missing information", "removed", "failed")):
                return "coral"
            return "amber"   # out of stock / missing offer / inactive / not confirmed

        st.caption(f"✅ {n_ok} confirmed created · ⏳ {len(uncreated)} not confirmed. "
                   f"Run the health-check to fill **Status** with Amazon's real state "
                   f"(Active / Out of stock / Missing offer / Missing information / "
                   f"Search suppressed) and the **Issues** column.")

        rows_data = [{
            "Created": _when(r), "SKU": r["sku"], "Title": r["title"],
            "Price": f"AED {r['price']:,.0f}",
            "Type": (r["payload"].get("product_type", "—")
                     if isinstance(r["payload"], dict) else "—"),
            "ASIN": (str(r.get("asin") or "").strip() or "—"),
            "Status": _status_cell(r),
            "Issues": _issue_cell(r["sku"])}
            for r in ready]
        qdf = pd.DataFrame(rows_data)
        # Build a colour for each distinct Status value present.
        status_badges = {s: (s, _status_kind(s)) for s in {row["Status"] for row in rows_data}}
        # highlight predicates receive the DataFrame row → key off the Status column.
        styled_table(qdf, highlight={
            "row-good": lambda row: _status_kind(row["Status"]) == "green"
            and not str(row["Issues"]).startswith("⚠"),
            "row-warn": lambda row: _status_kind(row["Status"]) != "green"
            or str(row["Issues"]).startswith("⚠")},
            badge_cols={"Status": status_badges})
        export_buttons(qdf, "ready_to_list")

        # ── Health check: ask Amazon what each listing is missing, fix status ──
        bc1, bc2 = st.columns(2)
        if bc1.button("🩺 Re-check & health-check ALL on Amazon", use_container_width=True,
                      help="Re-confirm each SKU and pull Amazon's missing/incomplete-info issues."):
            new_issues, new_status, promoted = {}, {}, 0
            seen = []
            with st.spinner("Asking Amazon for each listing's status & issues…"):
                for r in ready:
                    sku = r["sku"]
                    if sku in seen:
                        continue
                    seen.append(sku)
                    try:
                        info = client().confirm_listing(sku)
                    except Exception as e:
                        info = {"exists": False, "amazon_status": "Check failed",
                                "all_issues": [{"severity": "ERROR",
                                "message": f"check failed: {e}", "attributes": []}]}
                    if info.get("exists") and info.get("asin"):
                        db.upsert_catalog_item(
                            sku=sku, asin=info["asin"], title=r.get("title", ""),
                            category=(r["payload"].get("product_type", "")
                                      if isinstance(r["payload"], dict) else ""),
                            status="listed")
                        promoted += 1
                    new_issues[sku] = info.get("all_issues", [])
                    new_status[sku] = (info.get("amazon_status")
                                       or ("Not on Amazon" if not info.get("exists") else ""))
            st.session_state["aic_issues"] = new_issues
            st.session_state["aic_status_map"] = new_status
            incomplete = sum(1 for v in new_issues.values() if v)
            st.success(f"Health-check done — {promoted} confirmed created, "
                       f"{incomplete} listing(s) missing/incomplete info (see below).")
            st.rerun()
        if uncreated and bc2.button(f"🗑️ Remove {len(uncreated)} not-created from list",
                                    use_container_width=True):
            n = db.delete_ready_to_list([r["sku"] for r in uncreated])
            st.session_state.pop("aic_issues", None)
            st.session_state.pop("aic_status_map", None)
            st.success(f"Removed {n} not-created entr(y/ies).")
            st.rerun()

        # ── Per-SKU "what's missing" detail, straight from Amazon ──
        flagged = {s: v for s, v in issues_map.items() if v}
        if flagged:
            st.markdown(alert(
                f"{len(flagged)} listing(s) are live but missing/incomplete info on Amazon. "
                "Open the editor above, fill the named field(s), and push again.",
                kind="amber", icon="⚠️"), unsafe_allow_html=True)
            with st.expander(f"⚠️ What's missing on {len(flagged)} listing(s)", expanded=True):
                for sku, iss in flagged.items():
                    st.markdown(f"**{sku}**")
                    for i in iss:
                        attrs = ", ".join(i.get("attributes") or [])
                        sev = "⛔" if i.get("severity") == "ERROR" else "⚠"
                        st.markdown(f"- {sev} {i.get('message', '')}"
                                    + (f"  _(field: `{attrs}`)_" if attrs else ""))
    else:
        st.caption("Nothing pushed yet.")


def render(nav=None) -> None:
    page_header("Inventory & Listing Intake",
                "Classify arrivals, plan restocks, and auto-create listings", icon="📥")
    t1, t2, t3 = st.tabs(["✨ New Arrivals", "🚚 Restock Pending", "🤖 Auto Item Creation"])
    with t1:
        _new_arrivals_tab()
    with t2:
        _restock_tab()
    with t3:
        _auto_creation_tab()
